"""
A script for using the run.py module as a subprocess to solve multiple Energy Hubs while changing a given parameter.

This functionality is now also provided by the Besos EvaluatorEH.
"""

import subprocess
import openpyxl as op
import numpy as np

LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

# Step 1: set the start and stop point of changing parameter and number of steps(+1) in line 18 to 20.
# Step 2: Set the sheet name and cell number of parameter in line 23, 24.
# Step 3: In line 22 (in_file), write the directory of the input excel.
# Step 4: In line 25 (model_out_dir), put the output directory which want to save in.
# Step 5: line 26 is the directory that the temporary file saves in. It doesn't need to change it.
# Step 6: The summary output file which contains total cost and carbon will save in line 29.
# You can change its name here.
# Step 7: in line 40, you can change the name of output file per each run.


def main():
    start = 671
    stop = 1
    steps = 11

    # in_file = 'Individual_Hubs/LI_Battery/PV_50/Diesel_27/Input_LI_PV50_D27.xlsx'
    in_file = 'Individual_Hubs/test.xlsx'
    sheet_name = 'Storages'
    cell_to_set = 'B4'
    model_out_dir = 'Individual_Hubs/LI_Battery/PV_50/Diesel_27'
    model_inputs = 'Individual_Hubs/temp_inputs.xls'
    sheet_to_read = 'Other'
    cells_to_read = ['B168', 'B170']
    final_output_file = 'Individual_Hubs/LI_Battery/PV_50/Diesel_27/output_LI_PV50_D27.xlsx'
    result_wb = op.Workbook()
    result_sheet = result_wb['Sheet']
    output_cells = ['Input value'] + cells_to_read
    for column, cell in zip(LETTERS, output_cells):
        result_sheet[f'{column}1'].value = cell

    for row,value in enumerate(np.linspace(start, stop, steps), start=2):
        wb = op.load_workbook(in_file)
        wb[sheet_name][cell_to_set].value = value
        wb.save(model_inputs)
        model_out_file = fr'{model_out_dir}/LI_PV50_D27_{int(value)}.xlsx'
        #the input file for run.py is specified in config.yaml
        print('starting subprocess')
        subprocess.run(['python', 'run.py', '--output', 'model_out_file',])
        print('file should be there')

        wb = op.load_workbook(model_out_file)
        result_sheet[f'A{row}'].value = value
        for col,cell_to_read in zip(LETTERS[1:], cells_to_read):
            print(f'writing {wb[sheet_to_read][cell_to_read].value} to {col}{row} for {value}')
            result_sheet[f'{col}{row}'].value = wb[sheet_to_read][cell_to_read].value
    result_wb.save(final_output_file)




if __name__ == '__main__':
    main()