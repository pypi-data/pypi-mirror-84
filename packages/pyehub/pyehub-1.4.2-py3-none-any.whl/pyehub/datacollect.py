"""
A script for printing out the contents of an excel spreadsheet.
"""

import openpyxl
from openpyxl import load_workbook

def main():
    """Main function."""
    book = openpyxl.load_workbook('Individual_Hubs/LI_Batresuls_11.xlsx')
    print(book.get_sheet_names())
    sheet = book.get_sheet_by_name('Other')

    # No of written Rows in sheet
    num_row = sheet.max_row
    # No of written Columns in sheet
    num_column = sheet.max_column
    # Reading each cell in excel
    for i in range(1, num_row+1):
        for j in range(1, num_column+1):
            print(sheet.cell(row=i, column=j).value)

if __name__ == '__main__':
    main()
