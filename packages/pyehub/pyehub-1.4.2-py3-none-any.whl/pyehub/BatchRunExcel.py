"""
A script used to solve a batch series of energy hubs in a combined PyEHub model.

This functionality is now also provided by the multiple hubs module.
"""

import argparse
import os
import re
import shutil
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import multiple_hubs

""" File naming convention:

In order for an excel file to be read as input correctly, it must
follow the follwing naming convention:

[any valid file name]_[id to be assosiated with it].xlsx

note,
	carbon/cost is there to allow you to specify whether a test is trying to minimize
	either cost or carbon.

"""

def get_output_fields(filenames: list, directory: str = None):
	""" get_output_fields(list, str)

		Gets the required outputs from the excel files passed in the
		list of filenames.

		Args:
			filenames(list): A list of the excel files to find outputs for.

			directory(str): The optional dicretory that these filenames are located at.
					for use only if the filenames do not include their directory.

		Returns:
			A list with every output requested from every file (duplicates are deleted.)
	"""

	#: Create the list to hold the values in.
	master_field_list = []

	for filename in filenames:

		#: Open the excel file for this loop.
		if directory is not None:
			book = openpyxl.load_workbook(directory + filename)
		else:
			book = openpyxl.load_workbook(filename)

		#: Open the capacities sheet which is set to have all of
		#: the required outputs in it.
		sheet = book["Capacities"]

		#: Loop through the header (skipping not output fields)\
		#: and get all of the required outputs. These exist from
		#: (row 1, col 2) to (row 1, final col)
		fields = sheet.iter_rows(min_row=1, max_row=1, min_col=2)

		#: Convert the generator to a list, and get the only row.
		#: then loop through each cell in the row, and append the
		#: value in that cell to the master list.
		for cell in list(fields)[0]:
			master_field_list.append(cell.value)

	#: Append total_cost and total_cost to the output list so that these
	#: values are automatically included in the output.
	master_field_list.append("total_cost")
	master_field_list.append("total_carbon")
	#: Convert the list to a set to remove all duplicates,
	#: then convert the set back to a list, and return.
	return_list = []
	for item in master_field_list:
		if item is not None:
			return_list.append(item)

	return sorted(list(set(return_list)))


def output_to_excel(outputs: dict, sheet, output_dict: dict):
	""" output_to_excel(dict, openpyxl.sheet, dict)

		Takes the output dictionary produced by multiple hubs, and
		writes specific fields to the passed excel sheet based on
		the output dictionary passed.

		Args:
			outputs(dict): The dictionary produced by multiple_hubs.

			sheet(openpyxl.sheet): The excel sheet to be written to.

			output_dict(dict): A dictionary that links keys found in the
					   outputs dict, and the column that they should
					   appear in, within the sheet.

	"""

	#: Get all of the columns that have content in them, and use
	#: that to figure out where the next blank column is.
	cols = list(sheet.iter_cols(min_col=1, max_col=1))[0]
	first_empty_row = len(cols) + 1


	#: Merge the cells that have large words togther. These are the
	#: first 4 cells appearing in any output (Filename, funcion,
	#: fjob, and hub).
	#: File name header cell
	sheet.merge_cells("A{}:B{}".format(first_empty_row, first_empty_row))
	#: Function header cell
	sheet.merge_cells("C{}:D{}".format(first_empty_row, first_empty_row))
	#: fjob header cell
	sheet.merge_cells("E{}:F{}".format(first_empty_row, first_empty_row))
	#: Hub header cell
	sheet.merge_cells("G{}:H{}".format(first_empty_row, first_empty_row))

	#: for each key and value pair in the formatting dictionary passed.
	for key, value in output_dict.items():

		#: If outputs[key] exists (in other words if the current key from output_dict,
		#: represents an output that is in this set of outputs, then set it to the corrisponding
		#: column based off the value of this item.
		try:
			sheet['{}{}'.format(value, first_empty_row)] = outputs[key]

		#: Otherwise move on to the next loop, as this output was not outputted by this file.
		except:
			continue

def main(batch_run_location: str="excel_files/batch_input/", output_directory: str="batchrun_outputs/", epsilon_n: int = 0):

	#: If argparse got no -d switch
	if output_directory is None:

		#:Set output directory to its default setting.
		output_directory = "batchrun_outputs/"

	#: Otherwise, if the user forgot to / or \ terminate their directory, then
	#: do it for them.
	elif output_directory[-1] != "/" and output_directory[-1] != "\\":
		output_directory = output_directory + "/"

	#: If argparse got no -o switch
	if batch_run_location is None:

		#: Set batch_run_location to its default setting.
		batch_run_location = "excel_files/batch_input/"

	#: Otherwise if the user forgot to / or \ terminate their directory, then
	#: do it for them.
	elif batch_run_location[-1] != "/" and batch_run_location[-1] != "\\":
		batch_run_location = batch_run_location + "/"

	#: If argparse got no -n switch
	if epsilon_n is None:

		#: Set epsilon_n to iss default setting.
		epsilon_n = 0

	#: If the output directory is not currently a directory, make it.
	if not os.path.exists(output_directory):
		os.makedirs(output_directory)

	#: Get a list of all files in the run location that end in .xlsx (excel's file
	#: extension)
	excel_files = [file for file in os.listdir(batch_run_location) if file[-5:] == ".xlsx" and file[0] != '$']

	file_data = [] #: An intermediate container to hold some of the information
		       #: That will be gathered and held on the files.

	#: Get the id of every file found in the run location,
	#: then store a tuple with (filename, fileid) into file_data.
	for file in excel_files:

		#: Initial position while looking for the _.
		_position = 0

		#: For each character in the file name (backwards),
		#: if the character is an underscore then we know that
		#: this is the character before the id starts.
		for character in file[::-1]:

			#: Check if the character is an _, if so then
			#: we're done looking and should break from the
			#: inner loop.
			if character == "_":
				break

			#: Otherwise subtract 1 from position.
			_position -= 1

		#: Append a tuple defined by (filename, fileid).
		#: file[_position:-5] gets the id because _position
		#: is the location of the LAST _ in the filename (ie, the
		#: one that is required to preceed the id) and there are five
		#: characters at the end of each file ".xlsx". Thus the range
		#: between _position, and 5 is the id (no matter how many numbers.)
		file_data.append((file, file[_position:-5]))

	#: Sort file data based off of the file ids so that
	#: all pairs of files are beside eachother in the list.
	file_data = sorted(file_data, key=lambda x: x[1])

	file_sets = [] #: The final container to hold tuples of pairs of filenames.

	#: Append a tuple defined by (filename1, filename1's pair) to file_sets.
	#: for each pair in file_data.
	for index, file in enumerate(file_data[::2]):
		file_sets.append((file, file_data[(index * 2) + 1]))

	#: Get all the required outputs from the found excel files.
	all_fields = get_output_fields(excel_files, directory=batch_run_location)

	#: The first column to put data in is alwasy I
	first_letter = 'I'

	#: The number of fields automatically appeneded after the capacities.
	#: These fields are:
	#:	total_carbon
	#:	total_cost
	extra_fields = 2

	#: Calculate the letter(s) of the final column to put data in. - 1 is to
	#: account for an off by one error. The complexity in these functions is
	#: to account for the ability of the sheet to be over 26 columns long. In
	#: that case, excel starts using AA, then AB, etc so this must be accounted for.
	final_letter = str(first_letter)
	for _ in range(len(all_fields) - 1 - extra_fields):
		#: If the final letter is not a Z, then just increase it by a letter.
		if final_letter[-1] != 'Z':
			final_letter = final_letter[:-1] + chr(ord(final_letter[-1]) + 1)
		#: Otherwise remove the Z and replace it with AA to account for
		#: excel's column naming.
		else:
			final_letter = final_letter[:-1] + "AA"

	#: Create a list of the extended row's letter(s) representing their columns.
	extended = [final_letter]
	for _ in range(extra_fields):
		if extended[-1][-1] != 'Z':
			extended.append(extended[-1][:-1] + chr(ord(extended[-1][-1]) + 1))
		else:
			extended.append(extended[-1][:-1] + "AA")

	#: Create an excel workbook, and open a sheet.
	output_book = Workbook()
	sheet = output_book.active

	#: Merge cells for proper formatting of headers.
	#: Filename header
	sheet.merge_cells("A1:B2")
	#: Function header
	sheet.merge_cells("C1:D2")
	#: fjob header
	sheet.merge_cells("E1:F2")
	#: Hub header
	sheet.merge_cells("G1:H2")
	#: Capacity header: must span from first output to last output.
	sheet.merge_cells("{}1:{}1".format(first_letter, final_letter))
	#: Totals header must span from final_letter + 1 to final_letter + 2
	sheet.merge_cells("{}1:{}1".format(extended[1], extended[2]))

	#: Assign header values to these cells.
	#: Filename header
	sheet['A1'] = "File Name"
	#: Function header
	sheet['C1'] = "Function"
	#: fjob header
	sheet['E1'] = "fjob"
	#: Hub header
	sheet['G1'] = "Hub"
	#: Use first_letter here instead of 'I' to increase readability.
	sheet['{}1'.format(first_letter)] = "Capacity"
	#: Totals header
	sheet['{}1'.format(extended[1])] = "Totals"

	#: A dictionary describing which column data for each header should
	#: go in. Used as the output_dict parameter for output_to_excel().
	pos_dict = {
		     #: header: Data column,
			"Filename": "A",
			"Function": "C",
			"fjob": "E",
			"Hub": "G"
		}
	#: Create a string to parse through the data columnsm
	#: and to set values with.
	fletter = str(first_letter)

	#: For each field, append that field's name as a subheader under
	#: capacity.
	for field in all_fields:

		#: Append the current field name to the column
		#: stored currently in fletter.
		sheet['{}2'.format(fletter)] = field

		#: Build the dictionary used in output_to_excel().
		#: set the key value of field to the current letter
		#: stored by fletter.
		pos_dict[field] = fletter

		#: Increment fletter by one. If fletter's last (only) character
		#: is Z, then change it to AA
		if fletter[-1] != 'Z':
			fletter = fletter[:-1] + chr(ord(fletter[-1]) + 1)
		else:
			fletter = fletter[:-1] + "AA"

	#: Centre the cells all of the cells that we have used as headers.
	#: The Filename header
	sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
	#: The function header
	sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
	#: The fjob header
	sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')
	#: The Hub header.
	sheet['G1'].alignment = Alignment(horizontal='center', vertical='center')
	#: Center the merged capacity cell.
	sheet['{}1'.format(first_letter)].alignment = Alignment(horizontal='center', vertical='center')
	#: Center the merged Totals cell.
	sheet['{}1'.format(extended[1])].alignment = Alignment(horizontal='center', vertical='center')

	#: Main execution loop. This handles the execution of each pair of files that was found.
	#: it also handles all of the outputting of the data found by the executions. Finally, it
	#: also handles looping over different carbon levels if epsilon_n is not 0.
	for fjob, file_set in enumerate(file_sets):

		#: Load the data from the first two excel files into the hub1 and hub2 files as required
		#: multiple_hubs.py
		shutil.copyfile(batch_run_location + file_set[0][0], 'docs/tutorials/network/hub1.xlsx')
		shutil.copyfile(batch_run_location + file_set[1][0], 'docs/tutorials/network/hub2.xlsx')

		#: Output current action as some actions may take a while and we want the user to know
		#: that we're not just frozen.
		print("Finding minimum cost for {} and {}".format(file_set[0][0][:-5], file_set[1][0][:-5]))

		#: Attempt to minimize the cost without worrying about the carbon output. Store the
		#: optimized data into cost_min.
		cost_min = multiple_hubs(False, output_directory + file_set[0][0][:_position - 1])

		#: Get which outputs are given by hub1 and hub2. NOTE these aren't the outputs,
		#: they are just the fields of important outputs to be logged in final_results.xlsx
		hub1_output_fields = get_output_fields([file_set[0][0]], directory=batch_run_location)
		hub2_output_fields = get_output_fields([file_set[1][0]], directory=batch_run_location)

		#: Create the base dictionary to output data with. This dictionary will
		#: be used as the outputs variable for output_to_excel().
		#: This base dictionary creates the information that isnt returned by
		#: multiple_hubs; this is mostly just readability data (ie used to distinguish
		#: between different runs.)
		hub1_outputs = {
				"Filename": file_set[0][0][:-5],
				"Function": "Minimize Cost",
				"fjob": (2 * fjob) + 1,
				"Hub": 1
			}

		#: Create the base dictionary to output data with. This dictionary will
		#: be used as the outputs variable for output_to_excel().
		#: This base dictionary creates the information that isnt returned by
		#: multiple_hubs; this is mostly just readability data (ie used to distinguish
		#: between different runs.)
		hub2_outputs = {
				"Filename": file_set[1][0][:-5],
				"Function": "Minimize Cost",
				"fjob": (2 * fjob) + 1,
				"Hub": 2
			}

		#: For each field that we need to store for hub 1,
		#: add an entry to hub1_outputs where the key is
		#: field, and the value is the value for that field
		#: found by running multiple_hubs()
		for field in hub1_output_fields:
			hub1_outputs[field] = cost_min[0][field]

		#: For each field that we need to store for hub 2,
		#: add an entry to hub1_outputs where the key is
		#: field, and the value is the value for that field
		#: found by running multiple_hubs()
		for field in hub2_output_fields:
			hub2_outputs[field] = cost_min[1][field]

		#: Output all of the data found to the working excel file.
		output_to_excel(outputs=hub1_outputs, sheet=sheet, output_dict=pos_dict)
		output_to_excel(outputs=hub2_outputs, sheet=sheet, output_dict=pos_dict)

		#: Find the average carbon used on a run that minimizes cost (and theoretically
		#: maximizes carbon.)
		avg_carbon = (cost_min[0]["total_carbon"] + cost_min[1]["total_carbon"]) / 2

		#: Find the value of the carbon that should be used as a maximum when
		#: multiplied by the current step.
		carbon_per_step = 0 if epsilon_n == 0 else avg_carbon / epsilon_n

		#: Handles the epsilon function.
		#: If epsilon_n is 0, then this will only run once in an attempt to
		#: minimize cost at certain (n * carbon_per_step) carbon maximums.
		#Otherwise, we will run espilon_n times + 1 for the minimize carbon run.
		for n in range(epsilon_n + 1):

			#: Print which files and what n value we're currently at (also print
			#: the maximum carbon for this run.)
			print("Running {} and {}  at n = {}, max carbon = {}"\
				.format(file_set[0][0][:-5], file_set[1][0][:-5], n, carbon_per_step * n))

			#: Attempt to minimize the cost, with a maximum carbon value of carbon_per_step * n.
			#: Note,
			#:	If this is the first/only loop then carbon will be maxed at zero, while attempting
			#:	to optimize for cost. This will simulate a no-carbon building, and will theoretically
			#:	be the maximum cost. On all other runs, the carbon will be maxed at carbon_per_step * n.
			carbon_min = multiple_hubs(True, output_directory + "n-{}".format(n) + file_set[0][0][:_position - 1],
						   carbon_per_step * n)

			#: Create the base dictionary to output data with. This dictionary will
			#: be used as the outputs variable for output_to_excel().
			#: This base dictionary creates the information that isnt returned by
			#: multiple_hubs; this is mostly just readability data (ie used to distinguish
			#: between different runs.)
			hub1_outputs = {
					"Filename": file_set[0][0][:-5],
					"Function": "Minimize Carbon; n = {}".format(n),
					"fjob": (fjob + 1) * 2,
					"Hub": 1
				}

			#: Create the base dictionary to output data with. This dictionary will
			#: be used as the outputs variable for output_to_excel().
			#: This base dictionary creates the information that isnt returned by
			#: multiple_hubs; this is mostly just readability data (ie used to distinguish
			#: between different runs.)
			hub2_outputs = {
					"Filename": file_set[1][0][:-5],
					"Function": "Minimize Carbon; n = {}".format(n),
					"fjob": (fjob + 1) * 2,
					"Hub": 2
				}

			#: For each field that we need to store for hub 1,
			#: add an entry to hub1_outputs where the key is
			#: field, and the value is the value for that field
			#: found by running multiple_hubs()
			for field in hub1_output_fields:
				hub1_outputs[field] = carbon_min[0][field]

			#: For each field that we need to store for hub 2,
			#: add an entry to hub1_outputs where the key is
			#: field, and the value is the value for that field
			#: found by running multiple_hubs()
			for field in hub2_output_fields:
				hub2_outputs[field] = carbon_min[1][field]

			#: Output all of the data found to the working excel file.
			output_to_excel(outputs=hub1_outputs, sheet=sheet, output_dict=pos_dict)
			output_to_excel(outputs=hub2_outputs, sheet=sheet, output_dict=pos_dict)

	#: Save the excel spread sheet that was used to store all of the runs's important outputs.
	output_book.save(output_directory + "final_results.xlsx")

if __name__ == "__main__":

	#: Create the argument parser to parse the passed arguments.
	parser = argparse.ArgumentParser(description="A commandline utility to batch run tests.")

	#: Add the -d argument.
	parser.add_argument('-d', '--directory_name', type=str, help='The directory containing the excel file to be run.\n'\
									'Default is excel_files/batch_input/')
	#: Add the -o argument.
	parser.add_argument('-o', '--output_name', type=str, help='The directory to store all outputted files in..\n'\
									'Default is batchrun_outputs/')
	#: Add the -n argument.
	parser.add_argument('-n', '--epsilon_n', type=int, help='The number of steps to break into for the epsilon constraint.')

	#: Parse the passed arguments.
	args = parser.parse_args()

	#: Pass the arguments that were passed.
	main(args.directory_name, args.output_name, args.epsilon_n)
