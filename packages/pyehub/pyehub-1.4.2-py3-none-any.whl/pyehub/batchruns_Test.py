"""
A script for testing the batchrun script's code.
"""

import subprocess
import openpyxl as op
import numpy as np

LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

# Step 1: set the start and stop point of changing parameter and number of steps(+1).
# Step 2: Set the sheet name and cell number of parameter (Sheet_name, cell_to_set).
# Step 3: In (in_file), write the directory of the input excel.
# Step 4: In (model_out_dir), put the output directory which want to save in.
# Step 5: model_inputs is the directory that the temporary file saves in. It doesn't need to change it.
# Step 6: The summary output file which contains total cost and carbon will save in 'final_output_file.
# You can change its name here.
# Step 7: in model_out_file, you can change the name of output file per each run.


def main():
    # This is the price ranges for diesel Fuel in $/kWh
    start_D = 0.1
    stop_D = 0.5
    steps_D = 9

    start_PV = 50
    stop_PV = 250
    steps_PV = 6

    in_file = 'Individual_Hubs/PV_Diesel/RUN2/Input_Diesel_PV_Tanklimit.xlsx'
    sheet_name = 'Streams'
    cell_to_set = 'D5'

    sheet_name_PV = 'Converters'
    cell_to_set_PV='B4'

    model_out_dir = 'Individual_Hubs/PV_Diesel/RUN2'
    model_inputs = 'Individual_Hubs/PV_Diesel/RUN2/temp_inputs.xls'
    sheet_to_read = 'Other'
    cells_to_read = ['B50', 'B52','B54','B60','B110','B168', 'B170']
    final_output_file = 'Individual_Hubs/PV_Diesel/RUN2/output_Diesel_Tanklimit.xlsx'
    result_wb = op.Workbook()
    result_sheet = result_wb['Sheet']
    output_cells = ['Input value'] + cells_to_read
    for column, cell in zip(LETTERS, output_cells):
        result_sheet[f'{column}1'].value = cell

    for roww,valuee in enumerate(np.linspace(start_PV, stop_PV, steps_PV), start=2):
        wb = op.load_workbook(in_file)
        wb[sheet_name_PV][cell_to_set_PV].value = valuee

        for row,value in enumerate(np.linspace(start_D, stop_D, steps_D), start=2):
            wb = op.load_workbook(in_file)
            wb[sheet_name][cell_to_set].value = value
            wb.save(model_inputs)
            model_out_file = fr'{model_out_dir}/PV{valuee}_D{value}_TL.xlsx'
            #the input file for run.py is specified in config.yaml
            subprocess.run(['python', 'run.py', '--output', model_out_file])


            wb = op.load_workbook(model_out_file)
            result_sheet[f'A{row}'].value = value
            for col,cell_to_read in zip(LETTERS[1:], cells_to_read):
                print(f'writing {wb[sheet_to_read][cell_to_read].value} to {col}{row} for {value}')
                result_sheet[f'{col}{row}'].value = wb[sheet_to_read][cell_to_read].value
        result_wb.save(final_output_file)




if __name__ == '__main__':
    main()
