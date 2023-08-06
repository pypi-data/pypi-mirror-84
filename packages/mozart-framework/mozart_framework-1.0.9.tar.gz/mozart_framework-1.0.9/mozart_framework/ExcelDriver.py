#=============================================================
#=================     Library Import    =====================
#=============================================================

import openpyxl
# Importa a parte de Subprocesso

import os

from pathlib import Path


#=============================================================
#=================        class          =====================
#=============================================================

# Classe Excel Resposável pela automatização do processo usando o Excel


class ExcelAutomator:

    # Application
    workbook = openpyxl.Workbook
    
    sheet = ""

    configDict = {}

    cellList = ""

    cell  = ""

    def __init__(self):
        nothing = None

    def createWorkbook(self,ExcelPath):
        self.workbook = openpyxl.Workbook()
        self.workbook.save(ExcelPath)

    def openWorkbook(self,ExcelPath):
        self.workbook = openpyxl.load_workbook(ExcelPath)

    def selectSheetByName(self, sheetName):
        self.sheet = self.workbook[sheetName]

    def readActualSheetToDictionary(self):
        for row in self.sheet.iter_rows():
            self.configDict[row[0].value] = row[1].value

    def readStandardConfigurationFile(self):
        self.openWorkbook(os.path.join(Path(os.path.dirname(os.path.abspath(__file__))).parent, "Config" , "Config.xlsx"))
        self.selectSheetByName("Settings")
        self.readActualSheetToDictionary()
        self.selectSheetByName("Assets")
        self.readActualSheetToDictionary()

        return self.configDict
        
