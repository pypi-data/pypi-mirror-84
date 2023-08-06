# i*- coding: utf-8 -*-
# * Authors:
#       * Arezki Feth <f.a@majerti.fr>;
#       * Miotte Julien <j.m@majerti.fr>;
#       * TJEBBES Gaston <g.t@majerti.fr>
"""
Excel exportation module
"""

import itertools
import logging
import openpyxl
from openpyxl.styles import (
    Color,
    Font,
)

import io
from string import ascii_uppercase

from sqla_inspect.export import (
    BaseExporter,
    SqlaExporter,
)
from sqla_inspect.base import Registry


log = logging.getLogger(__name__)


# A, B, C, ..., AA, AB, AC, ..., ZZ
ASCII_UPPERCASE = list(ascii_uppercase) + list(
    ''.join(duple)
    for duple in itertools.combinations_with_replacement(ascii_uppercase, 2)
    )

# To be overriden by end user
FORMAT_REGISTRY = Registry()


class XlsWriter(object):
    """
    Class providing common tools to write excel files from tabular datas

    Has to be subclassed, the subclass should provide a _datas and a headers
    attribute that contains the datas to render and the headers

        _datas

            list of tuples (each tuple is a row)

        headers

            list of dict containing the label of each column:
                {'label': <a label>}

    """
    title = u"Export"

    def __init__(self, guess_types=True, worksheet=None, **kw):
        if worksheet is None:
            self.book = openpyxl.workbook.Workbook(guess_types=guess_types)
            self.worksheet = self.book.active
            self.worksheet.title = self.title
        else:
            self.worksheet = worksheet
            self.book = worksheet.parent
        self.options = kw

    def save_book(self, f_buf=None):
        """
        Return a file buffer containing the resulting xls

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        if f_buf is None:
            f_buf = io.BytesIO()
        f_buf.write(openpyxl.writer.excel.save_virtual_workbook(self.book))
        f_buf.seek(0)
        return f_buf

    def set_color(self, cell, color):
        """
        Set the given color to the provided cell

            cell

                A xls cell object

            color

                A openpyxl color var
        """
        cell.style = cell.style.copy(font=Font(color=Color(rgb=color)))

    def format_row(self, row):
        """
        The render method expects rows as lists, here we switch our row format
        from dict to list respecting the order of the headers
        """
        res = []
        headers = getattr(self, 'headers', [])
        for column in headers:
            column_name = column['name']
            value = row.get(column_name, '')
            if hasattr(self, "format_%s" % column_name):
                value = getattr(self, "format_%s" % column_name)(value)
            res.append(value)
        return res

    def _populate(self):
        """
        Populate headers and rows before writing our book
        """
        self._render_headers()
        self._render_rows()

    def render(self, f_buf=None):
        """
        Definitely render the workbook

        :param obj f_buf: A file buffer supporting the write and seek
        methods
        """
        self._populate()

        return self.save_book(f_buf)

    def _render_rows(self):
        """
        Render the rows in the current stylesheet
        """
        _datas = getattr(self, '_datas', ())
        headers = getattr(self, 'headers', ())
        for index, row in enumerate(_datas):
            row_number = index + 2
            for col_num, value in enumerate(row):
                cell = self.worksheet.cell(row=row_number, column=col_num + 1)
                if value is not None:
                    cell.value = value
                else:
                    cell.value = ""
                if len(headers) > col_num:
                    header = headers[col_num]
                    format = get_cell_format(header)
                    if format is not None:
                        cell.number_format = format

    def _render_headers(self):
        """
        Write the headers row
        """
        headers = getattr(self, 'headers', ())
        for index, col in enumerate(headers):
            # We write the headers
            cell = self.worksheet.cell(row=1, column=index + 1)
            cell.value = col['label']

        index += 1

        extra_headers = getattr(self, 'extra_headers', ())
        for add_index, col in enumerate(extra_headers):
            cell = self.worksheet.cell(row=1, column=add_index + index + 1)
            cell.value = col['label']

    def set_title(self, title):
        self.worksheet.title = title

    def set_headers(self, headers):
        """
        Set the headers of our writer
        :param list headers: list of dict with at least a label key
        (label is mandatory : used for the export)

        Headers are filtered and ordered regarding the order option
        """
        self.headers = []
        if 'order' in self.options:
            for element in self.options['order']:
                for header in headers:
                    if header.get('key', header['label']) == element:
                        self.headers.append(header)
                        break
        else:
            self.headers = headers


def get_cell_format(column_dict, key=None):
    """
    Return the cell format for the given column

    :param column_dict: The column datas collected during inspection
    :param key: The exportation key
    """
    format = column_dict.get('format')
    prop = column_dict.get('__col__')

    if format is None and prop is not None:
        if hasattr(prop, 'columns'):
            sqla_column = prop.columns[0]
            column_type = getattr(sqla_column.type, 'impl', sqla_column.type)
            format = FORMAT_REGISTRY.get_item(column_type)
    return format


class SqlaXlsExporter(XlsWriter, SqlaExporter):
    """
    Main class used for exporting datas to the xls format

    Models attributes output can be customized through the info param :

        Column(Integer, infos={'export':
            {'excel':<excel_specific_options>,
             <main_export_options>
            }
        }

    main_export_options and excel_specific_options can be :

        label

            label of the column header

        format

            a function that will be fired on each row to format the output

        related_key

            If the attribute is a relationship, the value of the given attribute
            of the related object will be used to fill the cells

        exclude

            This data will not be inserted in the export if True

    Usage:

        a = SqlaXlsExporter(MyModel)
        for i in MyModel.query().filter(<myfilter>):
            a.add_row(i)
        a.render()
    """
    config_key = 'excel'

    def __init__(self, model, guess_types=True, worksheet=None, **kw):
        self.guess_types = guess_types
        self.is_root = worksheet is None
        XlsWriter.__init__(
            self,
            guess_types=guess_types,
            worksheet=worksheet,
            **kw
        )
        SqlaExporter.__init__(self, model=model, **kw)

    def _get_related_exporter(self, related_obj, column):
        """
        returns an SqlaXlsExporter for the given related object and stores it in
        the column object as a cache
        """
        result = column.get('sqla_xls_exporter')
        if result is None:
            worksheet = self.book.create_sheet(
                title=column.get('label', 'default title')
            )
            result = column['sqla_xls_exporter'] = SqlaXlsExporter(
                related_obj.__class__,
                worksheet=worksheet
            )
        return result

    def _get_relationship_cell_val(self, obj, column):
        """
        Return the value to insert in a relationship cell
        Handle the case of complex related datas we want to handle
        """
        val = SqlaExporter._get_relationship_cell_val(self, obj, column)
        if val == "":
            related_key = column.get('related_key', None)

            if column['__col__'].uselist and related_key is None \
                    and self.is_root:

                # on récupère les objets liés
                key = column['key']
                related_objects = getattr(obj, key, None)
                if not related_objects:
                    return ""
                else:
                    exporter = self._get_related_exporter(
                        related_objects[0],
                        column,
                    )
                    for rel_obj in related_objects:
                        exporter.add_row(rel_obj)

        return val

    def _populate(self):
        """
        Enhance the default populate script by handling related elements
        """
        XlsWriter._populate(self)
        for header in self.headers:
            if "sqla_xls_exporter" in header:
                header['sqla_xls_exporter']._populate()


class XlsExporter(XlsWriter, BaseExporter):
    """
    A main xls exportation tool (without sqlalchemy support)

    writer = MyXlsExporter()
    writer.add_row({'key': u'La valeur de la cellule de la colonne 1'})
    writer.render()
    """
    headers = ()

    def __init__(self, guess_types=True, **kw):
        XlsWriter.__init__(
            self,
            guess_types=guess_types,
            **kw
        )
        BaseExporter.__init__(self, **kw)
