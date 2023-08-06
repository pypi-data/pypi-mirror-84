# -*- coding: utf-8 -*-
"""Contains the classes for Functions objects.

There are three main types of Functions:
    1) SeparationFunction: separates the imported pandas DataFrame into
                           multiple DataFrames.
    2) CalculationFunction: performs a calculation on each of the entries.
    3) SummaryFunction: performs a calculation once per sample or once per
                        dataset.

@author: Donald Erb
Created on Jul 31, 2020

#TODO need to update all docstrings
"""


import itertools

from openpyxl.utils.cell import get_column_letter as _get_column_letter


class FunctionBase:
    """
    Base class for all other Function classes.

    Attributes
    ----------
    name : str
        The string representation for the Function.
    target_columns : str or (list, tuple)
        A string or list/tuple of strings designating the target columns
        for the Function.

    Parameters
    ----------
    name : str
        The string representation for the Function.
    target_columns : str or (list, tuple)
        A string or list/tuple of strings designating the target columns
        for the Function.

    """

    def __init__(self, name, target_columns):
        """
        Raises
        ------
        ValueError
            Raised if an empty string is given as the name.

        """

        if name:
            self.name = name
        else:
            raise ValueError('Function name cannot be a blank string.')

        if isinstance(target_columns, str):
            self.target_columns = [target_columns]
        else:
            self.target_columns = target_columns


    def __str__(self):
        return f'{self.__module__}.{self.__class__.__name__} {self.name}'


    def __repr__(self):
        return f'<{str(self)}>'



class SeparationFunction(FunctionBase):
    """
    Function used to separate a single dataframe into multiple dataframes.
    """

    def __init__(self, name, target_columns, function, function_kwargs=None):
        """


        Parameters
        ----------
        name : TYPE
            DESCRIPTION.
        target_columns : TYPE
            DESCRIPTION.
        function : TYPE
            DESCRIPTION.
        function_kwargs : TYPE, optional
            DESCRIPTION. The default is None.

        Returns
        -------
        None.

        """

        super().__init__(name, target_columns)
        self.function = function
        self.function_kwargs = function_kwargs if function_kwargs is not None else {}


    def separate_dataframes(self, dataset, column_reference):
        """


        Parameters
        ----------
        dataset : TYPE
            DESCRIPTION.
        column_reference : TYPE
            DESCRIPTION.

        Returns
        -------
        new_datasets : list
            DESCRIPTION
        new_column_referece : list
            DESCRIPTION

        """

        new_datasets = []
        new_column_referece = []
        for i, sample in enumerate(dataset):
            new_samples = []
            new_references = []
            for j, dataframe in enumerate(sample):

                target_columns = [
                    int(column_reference[i][j][f'index_{column}']) for column in self.target_columns
                ]

                new_dataframes = self.function(
                    dataframe, target_columns, **self.function_kwargs
                )

                for df in new_dataframes:
                    # ensures that the new dataframe indices start at 0
                    df.reset_index(drop=True, inplace=True)

                new_samples.extend(new_dataframes)
                new_references.extend([column_reference[i][j]] * len(new_dataframes))

            new_datasets.append(new_samples)
            new_column_referece.append(new_references)

        return new_datasets, new_column_referece


class CalculationFunction(FunctionBase):
    """
    Function that performs a calculation for every entry in each sample.
    """

    def __init__(self, name, target_columns, functions, added_columns=1,
                 function_kwargs=None):
        """


        Parameters
        ----------
        name : TYPE
            DESCRIPTION.
        target_columns : TYPE
            DESCRIPTION.
        functions : TYPE
            DESCRIPTION.
        added_columns : int or str or tuple(str)/list(str), optional
            DESCRIPTION. The default is 0.
        function_kwargs : TYPE, optional
            DESCRIPTION. The default is None.
        callback_functions : TYPE, optional
            DESCRIPTION. The default is None.

        Returns
        -------
        None.

        """

        super().__init__(name, target_columns)

        if not added_columns:
            raise ValueError(f'Added columns for "{self.name}" must be > 0 or a string.')
        if isinstance(added_columns, int):
            self.added_columns = added_columns
        elif isinstance(added_columns, str):
            self.added_columns = [added_columns]
        else:
            self.added_columns = added_columns
        #TODO maybe just ensure that the input is a list/tuple with len=2, and that each is a dictionary
        if not isinstance(functions, (list, tuple)):
            self.functions = (functions, functions)
        else:
            self.functions = functions

        if function_kwargs is None:
            self.function_kwargs = ({}, {})
        elif not isinstance(function_kwargs, (list, tuple)):
            self.function_kwargs = (function_kwargs, function_kwargs)
        else:
            self.function_kwargs = function_kwargs


    def do_function(self, dataset, reference, index, first_column, first_row):
        """


        Parameters
        ----------
        dataset : TYPE
            DESCRIPTION.
        reference : TYPE
            DESCRIPTION.
        index : int
            Either 0 or 1. If 0, do Excel formulas; if 1, do python formulas.
        first_column : int
            The first Excel column to use; corresponds to the actual column
            number in Excel (ie is 1-based rather than 0-based), so 1 denotes 'A'.
        first_row : int
            The first Excel row to use; corresponds to the actual row number
            in Excel (ie is 1-based rather than 0-based), so 1 denotes the
            Excel row 1 and is the first row in the Excel workbook.

        Returns
        -------
        dataset : pd.DataFrame
            The input dataframe modified by the function.

        """

        if index == 1:
            excel_columns = None
        else:
            excel_columns = [
                _get_column_letter(num) for num in range(first_column, len(dataset.columns) + first_column)
            ]

        target_columns = [reference[target] for target in self.target_columns]
        added_columns = reference[self.name]

        dataset = self.functions[index](
            dataset, target_columns, added_columns, excel_columns, first_row,
            **self.function_kwargs[index]
        )

        return dataset


class SummaryFunction(CalculationFunction):
    """
    Calculation that is only performed once per sample or once per dataset.
    """

    def __init__(self, name, target_columns, functions, added_columns=1,
                 function_kwargs=None, sample_summary=True):
        """


        Parameters
        ----------
        name : TYPE
            DESCRIPTION.
        target_columns : TYPE
            DESCRIPTION.
        functions : TYPE
            DESCRIPTION.
        added_columns : TYPE, optional
            DESCRIPTION. The default is 1.
        function_kwargs : TYPE, optional
            DESCRIPTION. The default is None.
        sample_summary : bool
            If True, denotes that the SummaryFunction summarizes a sample; if False,
            denotes that the SummaryFunction summarizes a dataset.

        """

        super().__init__(name, target_columns, functions, added_columns, function_kwargs)
        self.sample_summary = sample_summary
