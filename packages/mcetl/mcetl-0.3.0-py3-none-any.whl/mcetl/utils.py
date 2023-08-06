# -*- coding: utf-8 -*-
"""Provides utility functions, classes, and constants.

Useful functions are put here in order to prevent circular importing
within the other files.

@author: Donald Erb
Created on Jul 15, 2020

Attributes
----------
DEFAULT_FITTING_FORMATS : dict
    The default openpyxl styles to use when writing the peak fitting
    results to Excel.
PROCEED_COLOR : tuple(str, str)
    The button color for all buttons that proceed to the next window.

"""


import operator
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import PySimpleGUI as sg
import matplotlib.pyplot as plt


DEFAULT_FITTING_FORMATS = {
    'fitting_header_even': {
        'font': Font(size=12, bold=True),
        'fill': PatternFill(
            fill_type='solid', start_color='F9B381', end_color='F9B381'
        ),
        'border': Border(bottom=Side(style='thin')),
        'alignment': Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
    },
    'fitting_header_odd': {
        'font': Font(size=12, bold=True),
        'fill': PatternFill(
            fill_type='solid', start_color='73A2DB', end_color='73A2DB'
        ),
        'border': Border(bottom=Side(style='thin')),
        'alignment': Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
    },
    'fitting_subheader_even': {
        'font': Font(bold=True),
        'fill': PatternFill(
            fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'
        ),
        'border': Border(bottom=Side(style='thin')),
        'alignment': Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
    },
    'fitting_subheader_odd': {
        'font': Font(bold=True),
        'fill': PatternFill(
            fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'
        ),
        'border': Border(bottom=Side(style='thin')),
        'alignment': Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )
    },
    'fitting_columns_even': {
        'fill': PatternFill(
            fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'
        ),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'number_format': '0.00',
    },
    'fitting_columns_odd': {
        'fill': PatternFill(
            fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'
        ),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'number_format': '0.00',
    },
    'fitting_descriptors_even': {
        'font': Font(bold=True),
        'fill': PatternFill(
            fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'
        ),
        'alignment': Alignment(
            horizontal='center', vertical='center', wrap_text=True
        ),
        'number_format': '0.000',
    },
    'fitting_descriptors_odd': {
        'font': Font(bold=True),
        'fill': PatternFill(
            fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'
        ),
        'alignment': Alignment(
            horizontal='center', vertical='center', wrap_text=True
        ),
        'number_format': '0.000',
    }
}

PROCEED_COLOR = ('white', '#00A949')


class WindowCloseError(Exception):
    """Custom exception to allow exiting a GUI window to stop the program."""


def safely_close_window(window):
    """
    Closes a PySimpleGUI window and removes the window and its layout.

    Used when exiting a window early by manually closing the window. Ensures
    that the window is properly closed and then raises a
    WindowCloseError exception, which can be used to determine that the window
    was manually closed.

    Parameters
    ----------
    window : sg.Window
        The window that will be closed.

    Raises
    ------
    WindowCloseError
        Custom exception to notify that the window has been closed earlier
        than expected.

    """

    window.close()
    raise WindowCloseError('Window was closed earlier than expected.')


def string_to_unicode(input_list):
    r"""
    Converts strings to unicode by replacing '\\' with '\'.

    Necessary because the user input from PySimpleGui's InputText element
    will convert any '\' input by the user to '\\', which will not
    be converted to the desired unicode. If the string already has unicode
    characters, it will be left alone.

    Also converts things like '\\n' and '\\t' to '\n' and '\t', respectively,
    so that inputs are correctly interpreted.

    Parameters
    ----------
    input_list : (list, tuple) or str
        A container of strings or a single string.

    Returns
    -------
    output : (list, tuple) or str
        A container of strings or a single string, depending on the input,
        with the unicode correctly converted.

    Notes
    -----
    Uses raw_unicode_escape encoding to ensure that any existing unicode is
    correctly decoded; otherwise, it would translate incorrectly.

    If using mathtext in matplotlib and want to do something like $\nu$,
    input $\\nu$ in the gui, which gets converted to $\\\\nu$ by PySimpleGUI,
    and in turn will be converted back to $\nu$ by this fuction.

    """

    if isinstance(input_list, str):
        input_list = [input_list]
        return_list = False
    else:
        return_list = True

    output = []
    for entry in input_list:
        if '\\' in entry:
            entry = entry.encode('raw_unicode_escape').decode('unicode_escape')
        output.append(entry)

    return output if return_list else output[0]


def validate_inputs(window_values, integers=None, floats=None,
                    strings=None, user_inputs=None, constraints=None):
    """
    Validates entries from a PySimpleGUI window and converts to the desired type.

    Parameters
    ----------
    window_values : dict
        A dictionary of values from a PySimpleGUI window, generated by using
        window.read().
    integers : list, optional
        A list of lists (see Notes below), with each key corresponding
        to a key in the window_values dictionary and whose values should
        be integers.
    floats : list, optional
        A list of lists (see Notes below), with each key corresponding
        to a key in the window_values dictionary and whose values should
        be floats.
    strings : list, optional
        A list of lists (see Notes below), with each key corresponding
        to a key in the window_values dictionary and whose values should
        be non-empty strings.
    user_inputs : list, optional
        A list of lists (see Notes below), with each key corresponding
        to a key in the window_values dictionary and whose values should
        be a certain data type; the values are first determined by
        separating each value using ',' (default) or the last index.
    constraints : list, optional
        A list of lists (see Notes below), with each key corresponding
        to a key in the window_values dictionary and whose values should
        be ints or floats constrained between upper and lower bounds.

    Returns
    -------
    bool
        True if all data in the window_values dictionary is correct.
        False if there is any error with the values in the window_values dictionary.

    Notes
    -----
    Inputs for integers, floats, and strings are [[key, display text],].
    For example: [['peak_width', 'peak width']]

    Inputs for user_inputs are
        [[key, display text, data type, allow_empty_input (optional), separator (optional)],],
    where separator is a string, and allow_empty_input is a boolean.
    If no separator is given, it is assumed to be a comma (','), and
    if no allow_empty_input value is given, it is assumed to be False.
    user_inputs can also be used to run the inputs through a function by setting
    the data type to a custom function. Use None as the separator if only a
    single value is wanted.
    For example: [
        ['peak_width', 'peak width', float], # ensures each entry is a float
        ['peak_width_2', 'peak width 2', int, False, ';'], # uses ';' as the separator
        ['peak_width_3', 'peak width 3', function, False, None], # no separator, verify with function
        ['peak_width_4', 'peak width 4', function, True, None] # allows empty input
    ]

    Inputs for constraints are
        [[key, display text, lower bound, upper bound (optional)],],
    where lower and upper bounds are strings with the operator and bound, such
    as "> 10". If lower bound or upper bound is None, then the operator and
    bound is assumed to be >=, -np.inf and <=, np.inf, respectively.
    For example: [
        ['peak_width', 'peak width', '> 10', '< 20'], # 10 < peak_width < 20
        ['peak_width_2', 'peak width 2', None, '<= 5'] # -inf <= peak_width_2 <= 5
        ['peak_width_3', 'peak width 3', '> 1'] # 1 < peak_width_2 <= inf
    ]

    The display text will be the text that is shown to the user if the value
    of window_values[key] fails the validation.

    """

    if integers is not None:
        for entry in integers:
            try:
                window_values[entry[0]] = int(window_values[entry[0]])
            except:
                sg.popup(f'Need to enter integer in "{entry[1]}".\n', title='Error')
                return False

    if floats is not None:
        for entry in floats:
            try:
                window_values[entry[0]] = float(window_values[entry[0]])
            except:
                sg.popup(f'Need to enter number in "{entry[1]}".\n', title='Error')
                return False

    if strings is not None:
        for entry in strings:
            if not window_values[entry[0]]:
                sg.popup(f'Need to enter information in "{entry[1]}".\n',
                         title='Error')
                return False

    if user_inputs is not None:
        for entry in user_inputs:
            if len(entry) > 4:
                allow_empty_input = entry[3]
                separator = entry[4]
            elif len(entry) > 3:
                allow_empty_input = entry[3]
                separator = ','
            else:
                allow_empty_input = False
                separator = ','

            if separator is None:
                inputs = [window_values[entry[0]]] if window_values[entry[0]] else []
            else:
                inputs = [val.strip() for val in window_values[entry[0]].split(separator) if val]

            try:
                if inputs:
                    values = [entry[2](inpt) for inpt in inputs]
                    window_values[entry[0]] = values if separator is not None else values[0]
                elif allow_empty_input:
                    window_values[entry[0]] = [] if separator is not None else ''
                else:
                    raise ValueError('Entry must not be empty.')

            except Exception as e:
                sg.popup(
                    f'Need to correct entry for "{entry[1]}".\n\nError:\n    {repr(e)}\n',
                    title='Error')
                return False

    if constraints is not None:
        operators = {'>': operator.gt, '>=': operator.ge,
                     '<': operator.lt, '<=': operator.le}
        for entry in constraints:
            if entry[2] is not None:
                lower_key, lower_bound = entry[2].split(' ')
                lower_bound = float(lower_bound) if '.' in lower_bound else int(lower_bound)
            else:
                lower_key = '>='
                lower_bound = float('-inf')

            if len(entry) > 3 and entry[3] is not None:
                upper_key, upper_bound = entry[3].split(' ')
                upper_bound = float(upper_bound) if '.' in upper_bound else int(upper_bound)
            else:
                upper_key = '<='
                upper_bound = float('inf')

            if not (operators[lower_key](window_values[entry[0]], lower_bound)
                    and operators[upper_key](window_values[entry[0]], upper_bound)):
                sg.popup(
                    (f'"{entry[1]}" must be {lower_key} {lower_bound} and '
                     f'{upper_key} {upper_bound}.\n'),
                    title='Error'
                )
                return False

    return True


def show_dataframes(dataframes, title='Raw Data'):
    """
    Used to show data to help select the right columns or datasets from the data.

    Parameters
    ----------
    dataframes : list or pd.DataFrame
        Either (1) a pandas DataFrame, (2) a list of DataFrames, or (3) a list
        of lists of DataFrames. The layout of the window will depend on the
        input type.
    title : str, optional
        The title for the popup window.

    Returns
    -------
    window : sg.Window or None
        If no exceptions occur, a PySimpleGUI window will be returned; otherwise,
        None will be returned.

    """

    try:
        if isinstance(dataframes, pd.DataFrame):
            single_file = True
            dataframes = [[dataframes]]
        else:
            single_file = False

            if isinstance(dataframes[0], pd.DataFrame):
                single_dataset = True
                dataframes = [dataframes]
            else:
                single_dataset = False

        tabs = [[] for df_collection in dataframes]
        for i, df_collection in enumerate(dataframes):
            for j, dataframe in enumerate(df_collection):

                data = dataframe.values.tolist()
                if any(not isinstance(col, str) for col in dataframe.columns):
                    header_list = [f'Column {num}' for num in range(len(data[0]))]
                else:
                    header_list = dataframe.columns.tolist()

                tabs[i].append(
                    sg.Table(values=data, headings=header_list, key=f'table_{i}{j}',
                             auto_size_columns=True, vertical_scroll_only=False,
                             num_rows=min(25, len(data)))
                )

        if single_file:
            layout = [tabs[0]]
        else:
            datasets = []
            for i, tab_group in enumerate(tabs):
                datasets.append(
                    [sg.Tab(f'Entry {j + 1}', [[table]],
                            key=f'entry_{i}{j}') for j, table in enumerate(tab_group)]
                )

            if single_dataset:
                layout = [
                    [sg.TabGroup([datasets[0]],
                                 tab_background_color=sg.theme_background_color())]
                ]

            else:
                tab_groups = []
                for i, tab_group in enumerate(datasets):
                    tab_groups.append(
                        [sg.Tab(f'Dataset {i + 1}',
                                [[sg.TabGroup([tab_group],
                                              tab_background_color=sg.theme_background_color())]])]
                    )
                layout = [
                    [sg.TabGroup(tab_groups,
                                 tab_background_color=sg.theme_background_color())]
                ]

        window = sg.Window(title, layout, resizable=True)

    except Exception as e: #TODO do I still need this try-except block?
        sg.popup('Error reading file:\n    ' + repr(e) + '\n', title='Error')
        window = None

    return window


def optimize_memory(dataframe, convert_objects=False):
    """
    Optimizes dataframe memory usage by converting data types.

    Optimizes object dtypes by trying to convert to other dtypes,
    if the pandas version is greater than 1.0.0.
    Optimizes numerical dtypes by downcasting to the most appropriate dtype.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe to optimize.
    convert_objects : bool, optional
        If True, will attempt to convert columns with object dtype
        if the pandas version is >= 1.0.0.

    Returns
    -------
    optimized_df : pd.DataFrame
        The memory-optimized dataframe.

    Notes
    -----
    Only converts int and float numeric types, not numpy types like float64,
    float 32, int16, etc.

    convert_objects is needed because currently, when object columns
    are converted to a dtype of string, the row becomes a StringArray object,
    which does not have the tolist() method curently implemented
    (as of pandas version 1.0.5). openpyxl's dataframe_to_rows method
    uses each row's tolist() method to convert the dataframe into a
    generator of rows, so having a StringArray row without a tolist
    method causes an exception when using openpyxl's dataframe_to_rows.
    This could be alleviated by using dataframe.to_excel to write to
    Excel directly rather than using dataframe_to_rows, but using the
    dataframe_to_rows offers a significant speed increase (using openpyxl's
    method results in a speed increae of ~ 30% since the cells are only
    iterated over once. If using dataframe.to_excel and then formatting,
    it requires iterating over all cells twice). I would
    rather have a speed increase with the downside of more memory usage.
    The dtypes can be still converted to string after writing to Excel, though.

    """

    optimized_df = dataframe.copy()

    if convert_objects and int(pd.__version__.split('.')[0]) > 0:
        # attempts to convert object columns to other dtypes
        objects = dataframe.select_dtypes(['object'])
        if len(objects.columns) > 0:
            optimized_df[objects.columns] = objects.convert_dtypes(
                convert_integer=False
            )

    ints = dataframe.select_dtypes(include=['int'])
    if len(ints.columns) > 0:
        optimized_df[ints.columns] = ints.apply(
            pd.to_numeric, downcast='integer', errors='ignore'
        )

    floats = dataframe.select_dtypes(include=['float'])
    if len(floats.columns) > 0:
        optimized_df[floats.columns] = floats.apply(
            pd.to_numeric, downcast='float', errors='ignore'
        )

    return optimized_df


def raw_data_import(window_values, file, show_popup=True):
    """
    Used to import data from the specified file into pandas DataFrames.

    Also used to show how data will look after using certain import values.

    Parameters
    ----------
    window_values : dict
        A dictionary with keys 'row_start', 'row_end', columns', 'separator',
        and optionally 'sheet'.
    file : str:
        A string containing the path to the file to be imported.
    show_popup : bool
        If True, will display a popup window showing a table of the data.

    Returns
    -------
    dataframes : list or None
        A list of dataframes containing the data after importing if show_popup
        is False, otherwise returns None.

    Notes
    -----
    Optimizes the memory usage of the imported data before returning.

    """

    try:
        row_start = window_values['row_start']
        row_end = window_values['row_end']
        column_numbers = window_values['columns']
        if window_values['separator'].lower() not in ('', 'none'):
            separator = window_values['separator']
        else:
            separator = None

        #if separator is not None: #TODO check whether this is needed since tkinter gives a raw string from the input; regex should work automatically
        #    separator = string_to_unicode(separator)

        if file.endswith('.xlsx'):
            first_col = int(window_values['first_col'].split(' ')[-1])
            last_col = int(window_values['last_col'].split(' ')[-1]) + 1
            columns = list(range(first_col, last_col))
            repeat_unit = window_values['repeat_unit']

            total_dataframe = pd.read_excel(
                file, window_values['sheet'], None, skiprows=row_start,
                skipfooter=row_end, usecols=columns, convert_float=not show_popup
            )

            column_indices = [num + first_col for num in column_numbers]
            dataframes = []
            for num in range(max(1, len(total_dataframe.columns) // repeat_unit)):
                indices = [(num * repeat_unit) + elem for elem in column_indices]
                dataframes.append(total_dataframe[indices])

        else:
            dataframes = [
                pd.read_csv(
                    file, skiprows=row_start, skipfooter=row_end, header=None,
                    sep=separator, usecols=column_numbers, engine='python'
                )
            ]

        if not show_popup:
            for i, dataframe in enumerate(dataframes):
                dataframe.columns = list(range(len(dataframe.columns)))
                dataframes[i] = optimize_memory(dataframe)

        else:
            window_1_open = False
            if file.endswith('.xlsx') and len(dataframes) > 1:
                window_1_open = True
                window_1 = show_dataframes(total_dataframe, 'Total Raw Data')
                window_0 = show_dataframes(dataframes, 'Imported Datasets')
            else:
                window_0 = show_dataframes(dataframes[0], 'Imported Dataset')

            if window_0:
                window_0_open = True
                while window_0_open or window_1_open: #TODO use sg.read_windows once pysimplegui is updated rather than using read(100)

                    if window_1_open:
                        event_1 = window_1.read(100)[0]
                        if event_1 == sg.WIN_CLOSED:
                            window_1.close()
                            window_1_open = False

                    event_0 = window_0.read(100)[0]
                    if event_0 == sg.WIN_CLOSED:
                        window_0.close()
                        window_0_open = False

            del window_0
            if file.endswith('.xlsx') and len(dataframes) > 1:
                del window_1
            dataframes = None # to clean up memory, dataframe is not needed

        return dataframes

    except Exception as e:
        sg.popup('Error reading file:\n    ' + repr(e) + '\n', title='Error')


def select_file_gui(data_source=None, file=None):
    """
    GUI to select a file and input the necessary options to import its data.

    Parameters
    ----------
    data_source : DataSource, optional
        The DataSource object used for opening the file.
    file: str, optional
        A string containing the path to the file to be imported.

    Returns
    -------
    values : dict
        A dictionary containing the items necessary for importing data from
        the selected file.

    """

    # Default values for if there is no file specified
    default_inputs = {
        'row_start': 0 if data_source is None else data_source.start_row,
        'row_end': 0 if data_source is None else data_source.end_row,
        'separator': '' if data_source is None else data_source.separator,
        'columns': '0, 1' if data_source is None else ', '.join([
            str(elem) for elem in data_source.column_numbers
        ]),
        'total_indices': None if data_source is None else data_source.column_numbers,
        'variable_indices': None if data_source is None else dict(
            zip(data_source.unique_variables,
                data_source.unique_variable_indices)
        ),
        'sheets': [],
        'sheet': '',
        'excel_columns': [],
        'first_column': '',
        'last_column': '',
        'repeat_unit': '',
        'initial_separator': '',
        'initial_columns': '',
        'initial_row_start': '',
        'initial_row_end': '',
        'initial_total_indices': None if data_source is None else [''] * len(data_source.column_numbers),
    }

    validations = {
        'integers': [['row_start', 'start row'], ['row_end', 'end row']],
        'user_inputs': [['columns', 'data columns', int]],
        'constraints': [['row_start', 'start row', '>= 0'],
                        ['row_end', 'end row', '>= 0']]
    }

    disable_excel = True
    disable_other = True
    disable_bottom = True

    if file is not None:
        disable_bottom = False

        if not file.endswith('.xlsx'):
            disable_other = False
        else:
            disable_excel = False

            dataframes = pd.read_excel(file, None, None, convert_float=False)
            sheet_names = list(dataframes.keys())
            sheet_0_len = len(dataframes[sheet_names[0]].columns)

            default_inputs.update({
                'sheets': sheet_names,
                'sheet': sheet_names[0],
                'excel_columns': [f'Column {num}' for num in range(sheet_0_len)],
                'first_column': 'Column 0',
                'last_column': f'Column {sheet_0_len - 1}',
                'repeat_unit': sheet_0_len,
                'separator': '',
                'columns': ', '.join(str(num) for num in range(sheet_0_len)),
                'row_start': 0,
                'row_end': 0,
                'total_indices': list(range(sheet_0_len)),
            })

            validations['integers'].append(
                ['repeat_unit', 'number of columns per dataset']
            )
            validations['constraints'].append(
                ['repeat_unit', 'number of columns per dataset', '> 0']
            )

        default_inputs.update({
            'initial_separator': default_inputs['separator'],
            'initial_columns': default_inputs['columns'],
            'initial_row_start': default_inputs['row_start'],
            'initial_row_end': default_inputs['row_end'],
            'initial_total_indices': default_inputs['total_indices'],
        })

    layout = [
        [sg.Text('Excel Workbook Options', relief='ridge', size=(38, 1),
                 justification='center', pad=(0, (15, 10)))],
        [sg.Text('Sheet to use:'),
         sg.Combo(default_inputs['sheets'], size=(17, 4), key='sheet',
                  default_value=default_inputs['sheet'], disabled=disable_excel,
                  readonly=True, enable_events=True)],
        [sg.Text('First Column:'),
         sg.Combo(default_inputs['excel_columns'], size=(17, 4),
                  key='first_col', readonly=True,
                  default_value=default_inputs['first_column'],
                  disabled=disable_excel, enable_events=True)],
        [sg.Text('Last Column:'),
         sg.Combo(default_inputs['excel_columns'], size=(17, 4), key='last_col',
                  readonly=True, default_value=default_inputs['last_column'],
                  disabled=disable_excel, enable_events=True)],
        [sg.Text('Number of columns per dataset:'),
         sg.Input(default_inputs['repeat_unit'], key='repeat_unit',
                  do_not_clear=True, disabled=disable_excel,
                  size=(3, 1), enable_events=True)],
        [sg.Text('Other Filetype Options', relief='ridge', size=(38, 1),
                 justification='center', pad=(5, (25, 10)))],
        [sg.Text('Separator (eg. , or ;)', size=(20, 1)),
         sg.Input(default_inputs['initial_separator'], key='separator',
                  disabled=disable_other, do_not_clear=True, size=(5, 1))],
        [sg.Text('=' * 34, pad=(5, (10, 10)))],
        [sg.Text('Enter data columns,\n separated by commas:',
                 tooltip='Starts at 0'),
         sg.Input(default_inputs['initial_columns'], key='columns',
                  do_not_clear=True, tooltip='Starts at 0', size=(10, 1),
                  enable_events=True, disabled=disable_bottom)],
        [sg.Text('Start row:', tooltip='Starts at 0', size=(8, 1)),
         sg.Input(default_inputs['initial_row_start'], key='row_start',
                  do_not_clear=True, size=(5, 1), disabled=disable_bottom,
                  tooltip='Starts at 0')],
        [sg.Text('End row: ', tooltip='Counts up from bottom. Starts at 0',
                 size=(8, 1)),
         sg.Input(default_inputs['initial_row_end'], key='row_end',
                  do_not_clear=True, size=(5, 1), disabled=disable_bottom,
                  tooltip='Counts up from bottom. Starts at 0')]
    ]

    if file is None:
        layout.insert(
            0,
            [sg.InputText('Choose a file', key='file', enable_events=True,
                          disabled=True, size=(28, 1), pad=(5, (10, 5))),
             sg.FileBrowse(key='file_browse', target='file', pad=(5, (10, 5)),
                           file_types=(("All Files", "*.*"),
                                       ("Excel Workbook", "*.xlsx"),
                                       ("CSV", "*.csv"),
                                       ("Text Files", "*.txt")))]
        )
    if data_source is not None:
        for variable in data_source.unique_variables:
            layout.extend([
                [sg.Text(f'Column of {variable} data:'),
                 sg.Combo(default_inputs['initial_total_indices'],
                          default_inputs['initial_total_indices'][default_inputs['variable_indices'][variable]],
                          size=(3, 1), readonly=True,
                          key=f'index_{variable}', disabled=disable_bottom)]
            ])

    layout.extend([
        [sg.Button('Test Import', pad=(5, (15, 5))),
         sg.Button('Next', bind_return_key=True, pad=(5, (15, 5)),
                   button_color=PROCEED_COLOR)]
    ])

    window = sg.Window('Data Import', layout)
    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED:
            safely_close_window(window)

        elif event == 'file':
            if values['file'] == 'Choose a file':
                continue

            elif values['file'].endswith('xlsx'):
                dataframes = pd.read_excel(values['file'], None, None,
                                           convert_float=False)
                sheet_names = list(dataframes.keys())
                sheet_0_len = len(dataframes[sheet_names[0]].columns)

                window['sheet'].update(values=sheet_names, value=sheet_names[0],
                                       readonly=True)
                col_list = [f'Column {num}' for num in range(sheet_0_len)]
                window['first_col'].update(values=col_list, value=col_list[0],
                                           readonly=True)
                window['last_col'].update(values=col_list, value=col_list[-1],
                                          readonly=True)
                window['repeat_unit'].update(value=sheet_0_len, disabled=False)
                window['separator'].update(value='', disabled=True)
                window['columns'].update(
                    value=', '.join(str(num) for num in range(sheet_0_len)),
                    disabled=False
                )
                window['row_start'].update(value='0', disabled=False)
                window['row_end'].update(value='0', disabled=False)

                if not any('repeat_unit' in entry for entry in validations['integers']):
                    validations['integers'].append(
                        ['repeat_unit', 'number of columns per dataset']
                    )
                    validations['constraints'].append(
                        ['repeat_unit', 'number of columns per dataset', '> 0']
                    )

                if data_source is not None:
                    _assign_indices(
                        window, list(range(sheet_0_len)),
                        default_inputs['variable_indices']
                    )

            else:
                window['sheet'].update(values=[], value='', disabled=True)
                window['first_col'].update(values=[], value='', disabled=True)
                window['last_col'].update(values=[], value='', disabled=True)
                window['repeat_unit'].update(value='', disabled=True)
                window['separator'].update(value=default_inputs['separator'],
                                           disabled=False)
                window['columns'].update(value=default_inputs['columns'],
                                         disabled=False)
                window['row_start'].update(value=default_inputs['row_start'],
                                           disabled=False)
                window['row_end'].update(value=default_inputs['row_end'],
                                         disabled=False)

                for i, entry in enumerate(validations['integers']):
                    if 'repeat_unit' in entry:
                        del validations['integers'][i]
                        break
                for i, entry in enumerate(validations['constraints']):
                    if 'repeat_unit' in entry:
                        del validations['constraints'][i]
                        break

                if data_source is not None:
                    for variable in data_source.unique_variables:
                        window[f'index_{variable}'].update(
                            values=default_inputs['total_indices'], readonly=True,
                            set_to_index=default_inputs['variable_indices'][variable]
                        )

        elif event == 'sheet':
            dataframe = dataframes[values['sheet']]
            window['repeat_unit'].update(value=len(dataframe.columns))
            cols = [f'Column {num}' for num in range(len(dataframe.columns))]
            window['first_col'].update(values=cols, value=cols[0])
            window['last_col'].update(values=cols, value=cols[-1])
            window['columns'].update(
                value=', '.join(str(i) for i in range(len(dataframe.columns)))
            )

            if data_source is not None:
                _assign_indices(
                    window, [num for num in range(len(dataframe.columns))],
                    default_inputs['variable_indices']
                )

        elif event in ('first_col', 'last_col'):
            first_col = int(values['first_col'].split(' ')[-1])
            last_col = int(values['last_col'].split(' ')[-1]) + 1

            if (values['repeat_unit']
                    and (last_col - first_col) < int(values['repeat_unit'])):

                new_len = last_col - first_col
                window['repeat_unit'].update(value=new_len)
                update_text = [num for num in range(new_len)]
                window['columns'].update(
                    value=', '.join(str(elem) for elem in update_text)
                )

                if data_source is not None:
                    _assign_indices(window, update_text,
                                    default_inputs['variable_indices'])

        elif event == 'repeat_unit' and values['repeat_unit']:
            try:
                update_text = [num for num in range(int(values['repeat_unit']))]
                window['columns'].update(
                    value=', '.join(str(elem) for elem in update_text)
                )

                if data_source is not None:
                    _assign_indices(window, update_text,
                                    default_inputs['variable_indices'])
            except ValueError:
                sg.popup('Please enter an integer in "number of columns per dataset"',
                         title='Error')

        elif event == 'columns':
            if data_source is not None:
                update_text = [
                    entry for entry in values['columns'].replace(' ', '').split(',') if entry
                ]
                _assign_indices(window, update_text,
                                default_inputs['variable_indices'])

        elif event in ('Next', 'Test Import'):
            if file is None and values['file'] == 'Choose a file':
                sg.popup('Please choose a file', title='Error')
                continue

            elif validate_inputs(values, **validations):
                if event == 'Test Import':
                    test_file = file if file is not None else values['file']
                    raw_data_import(values, test_file)
                else:
                    break

    window.close()
    del window

    if data_source is not None: # converts column numbers back to indices
        for key in (key for key in values if key.startswith('index_')):
            for i, col_num in enumerate(values['columns']):
                if int(values[key]) == col_num:
                    values[key] = i
                    break

    return values


def _assign_indices(window, columns, variables):
    """
    Updates the indices for each variable based on the column length.

    If there are more variables than available columns, the additional
    variables will all be assigned to the last value in columns.

    Parameters
    ----------
    window : sg.Window
        The PySimpleGUI window update.
    columns : list or tuple
        A list or tuple of column numbers.
    variables : dict
        A dictionary with variable names as keys and their target indices
        as values.

    Notes
    -----
    The updated element in the window is a sg.Combo element.

    """

    for variable in variables:
        if variables[variable] < len(columns):
            index = variables[variable]
        else:
            index = len(columns) - 1

        window[f'index_{variable}'].update(
            values=columns, set_to_index=index, readonly=True
        )


def open_multiple_files():
    """
    Creates a prompt to open multiple files and add their contents to a dataframe.

    Returns
    -------
    dataframes : list
        A list of dataframes containing the imported data from the selected
        files.

    """

    layout = [
        [sg.Text('Enter number of files to open:'),
         sg.Input('1', size=(10, 1), key='num_files')],
        [sg.Text('')],
        [sg.Button('Next', button_color=PROCEED_COLOR, bind_return_key=True)]
    ]

    validations = {
        'integers': [['num_files', 'number of files']],
        'constraints': [['num_files', 'number of files', '> 0']]
    }

    window = sg.Window('Get Files', layout)
    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED:
            num_files = False
            safely_close_window(window)
        else:
            if validate_inputs(values, **validations):
                num_files = values['num_files']
                break

    window.close()
    del window

    dataframes = []
    if num_files:
        try:
            for _ in range(num_files):
                import_values = select_file_gui()
                dataframes.extend(
                    raw_data_import(import_values, import_values['file'], False)
                )
        except (WindowCloseError, KeyboardInterrupt):
            pass

    return dataframes


def get_dpi_correction(dpi):
    """
    Calculates the correction factor needed to create a figure with the desired dpi.

    Necessary because some matplotlib backends (namely qt5Agg) will adjust
    the dpi of the figure after creation.

    Parameters
    ----------
    dpi : float or int
        The desired figure dpi.

    Returns
    -------
    dpi_correction : float
        The scaling factor needed to create a figure with the desired dpi.

    Notes
    -----
    The matplotlib dpi correction occurs when the operating system display
    scaling is set to any value not equal to 100% (at least on Windows,
    other operating systems are unknown). This may cause issues when
    using UHD monitors, but I cannot test.

    To get the desired dpi, simply create a figure with a dpi equal
    to dpi * dpi_correction.

    """

    with plt.rc_context({'interactive': False}):
        dpi_correction = dpi / plt.figure('dpi_corrrection', dpi=dpi).get_dpi()
        plt.close('dpi_corrrection')

    return dpi_correction


def save_excel_file(excel_writer):
    """
    Handles saving the Excel file and the various exceptions that can occur.

    Parameters
    ----------
    excel_writer : pd.ExcelWriter
        The pandas ExcelWriter object that contains all of the
        information about the Excel file being created.

    """

    # Ensures that the folder destination exist
    Path(excel_writer.path).parent.mkdir(parents=True, exist_ok=True)

    try_to_save = True
    while try_to_save:
        try:
            excel_writer.save()
            print('\nSaved Excel file.')
            break

        except PermissionError:
            try_to_save = sg.popup_ok(
                'Trying to overwrite Excel file. Please close the file.\n'
            )
