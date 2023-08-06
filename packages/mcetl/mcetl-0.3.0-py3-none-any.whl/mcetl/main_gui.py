# -*- coding: utf-8 -*-
"""Provides GUIs to import data depending on the data source used, process and/or fit the data, and save everything to Excel

@author: Donald Erb
Created on May 5, 2020

"""


import itertools
import json
from pathlib import Path
import sys
import traceback

from openpyxl.styles import NamedStyle
import pandas as pd
import PySimpleGUI as sg

from . import utils
from .datasource import DataSource
from .file_organizer import file_finder, file_mover
from .peak_fitting_gui import launch_peak_fitting_gui
from .plotting_gui import launch_plotting_gui


def _get_save_location():
    """
    Gets the correct filepath to save the previous_search.json depending on the operating system.

    Returns
    -------
    pathlib.Path
        The absolute path to where the previous_search.json file will
        be saved.

    """

    path = None
    if sys.platform.startswith('win'): # Windows
        path = Path('~/AppData/Local/mcetl')
    elif sys.platform.startswith('darwin'): # Mac
        path = Path('~/Library/Application Support/mcetl')
    elif sys.platform.startswith('linux'): # Linux
        path = Path('~/.config/mcetl')

    if path is None or not path.expanduser().parent.exists():
        # in case the Windows/Mac/Linux places are wrong
        path = Path('~/.mcetl')

    return path.expanduser()


def _write_to_excel(dataframes, data_source, labels,
                    excel_writer, plot_excel, plot_options):
    """
    Creates an Excel sheet from data within a list of dataframes.

    Parameters
    ----------
    dataframes : list(pd.DataFrame)
        A list of dataframes. Each dataframe contains all the raw data to
        put on one sheet in Excel.
    data_source : DataSource
        The selected DataSource.
    labels : list(dict)
        A list of dictionaries containing all of the sheet names, sample names,
        and subheader names. Each dictionary is for one dataset/Excel sheet.
        Relevant keys are 'sheet_name', 'sample_names', 'column_labels',
        'sample_summary_labels', 'dataset_summary_labels', and
        'summary_name'.
    excel_writer : pd.ExcelWriter
        The pandas ExcelWriter object that contains all of the
        information about the Excel file being created.
    plot_excel : bool
        If True, will create a simple plot in Excel using the data_source's
        x_plot_index and y_plot_index.
    plot_options : list(dict)
        A list of dictionaries with values used to create the Excel plot
        if plot_excel is True.

    """

    from openpyxl.utils.dataframe import dataframe_to_rows
    if plot_excel:
        from openpyxl.chart import Reference, Series, ScatterChart
        from openpyxl.chart.series import SeriesLabel, StrRef
        from openpyxl.utils.cell import get_column_letter

    # openpyxl uses 1-based indices
    first_row = data_source.excel_row_offset + 1
    first_column = data_source.excel_column_offset + 1

    for i, dataset in enumerate(dataframes):
        # Ensures that the sheet name is unique so it does not overwrite data;
        # not needed for openpyxl, but just a precaution
        current_sheets = [sheet.title.lower() for sheet in excel_writer.book.worksheets]
        sheet_name = labels[i]['sheet_name']
        sheet_base = sheet_name
        num = 1
        while True:
            if sheet_name.lower() not in current_sheets:
                break
            else:
                sheet_name = f'{sheet_base}_{num}'
                num += 1

        worksheet = excel_writer.book.create_sheet(sheet_name)

        # Header values and formatting
        for j, header in enumerate(labels[i]['sample_names'] + labels[i]['summary_name']):
            suffix = 'even' if j % 2 == 0 else 'odd'
            cell = worksheet.cell(
                row=first_row,
                column=first_column + sum(sum(entry) for entry in data_source.lengths[i][:j]),
                value=header
            )
            cell.style = 'header_' + suffix
            worksheet.merge_cells(
                start_row=first_row,
                start_column=first_column + sum(sum(entry) for entry in data_source.lengths[i][:j]),
                end_row=first_row,
                end_column=first_column + sum(sum(entry) for entry in data_source.lengths[i][:j + 1]) - 1
            )

        # Subheader values and formatting
        flattened_lengths = list(itertools.chain.from_iterable(data_source.lengths[i]))
        subheaders = iter(labels[i]['total_labels'])
        for j, entry in enumerate(flattened_lengths):
            suffix = 'even' if j % 2 == 0 else 'odd'
            for col_index in range(entry):
                cell = worksheet.cell(
                    row=first_row + 1,
                    column=first_column + col_index + sum(flattened_lengths[:j]),
                    value=next(subheaders)
                )
                cell.style = 'subheader_' + suffix

        # Dataset values and formatting
        rows = dataframe_to_rows(dataset, index=False, header=False)
        for row_index, row in enumerate(rows, first_row + 2):
            entry = 1
            suffix = 'even'
            cycle = itertools.cycle(['odd', 'even'])
            for column_index, value in enumerate(row, first_column):
                if (column_index + 1 - first_column) > sum(flattened_lengths[:entry]):
                    suffix = next(cycle)
                    entry += 1
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.style = 'columns_' + suffix

        worksheet.row_dimensions[first_row].height = 18
        worksheet.row_dimensions[first_row + 1].height = 30

        if plot_excel:
            x_col = plot_options[i]['x_plot_index']
            y_col = plot_options[i]['y_plot_index']
            x_min = plot_options[i]['x_min']
            x_max = plot_options[i]['x_max']
            y_min = plot_options[i]['y_min']
            y_max = plot_options[i]['y_max']
            last_row = len(dataset) + 1 + first_row
            index_modifier = -1 if labels[i]['sample_summary_labels'] else 0

            # Reverses x or y axes if min > max
            if None not in (x_min, x_max) and x_min > x_max:
                x_reverse = True
                x_min, x_max = x_max, x_min
            else:
                x_reverse = False

            if None not in (y_min, y_max) and y_min > y_max:
                y_reverse = True
                y_min, y_max = y_max, y_min
            else:
                y_reverse = False

            chart_attributes = {
                'title': plot_options[i]['chart_title'] if plot_options[i]['chart_title'] else None,
                'x_axis': {
                    'title': plot_options[i]['x_label'],
                    'crosses': 'max' if y_reverse else 'min',
                    'scaling': {
                        'min': x_min,
                        'max': x_max,
                        'orientation': 'maxMin' if x_reverse else 'minMax',
                        'logBase': 10 if plot_options[i]['x_log_scale'] else None
                    }
                },
                'y_axis': {
                    'title': plot_options[i]['y_label'],
                    'crosses': 'max' if x_reverse else 'min',
                    'scaling': {
                        'min': y_min,
                        'max': y_max,
                        'orientation': 'maxMin' if y_reverse else 'minMax',
                        'logBase': 10 if plot_options[i]['y_log_scale'] else None
                    }
                }
            }

            chart = ScatterChart()
            for key, attribute in chart_attributes.items():
                if not isinstance(attribute, dict):
                    setattr(chart, key, attribute)
                else:
                    for axis_attribute, value in attribute.items():
                        if not isinstance(value, dict):
                            setattr(getattr(chart, key), axis_attribute, value)
                        else:
                            for internal_attribute, internal_value in value.items():
                                setattr(
                                    getattr(getattr(chart, key), axis_attribute),
                                    internal_attribute, internal_value
                                )

            location = first_column
            for j in range(len(labels[i]['sample_names'])):
                for k in range(len(data_source.lengths[i][j]) + index_modifier):
                    series = Series(
                        Reference(
                            worksheet,
                            location + sum(data_source.lengths[i][j][:k]) + y_col,
                            first_row + 2,
                            location + sum(data_source.lengths[i][j][:k]) + y_col,
                            last_row
                        ),
                        xvalues=Reference(
                            worksheet,
                            location + sum(data_source.lengths[i][j][:k]) + x_col,
                            first_row + 2,
                            location + sum(data_source.lengths[i][j][:k]) + x_col,
                            last_row
                        )
                    )
                    series.title = SeriesLabel(
                        StrRef(f"'{sheet_name}'!{get_column_letter(location)}{first_row}")
                    )
                    chart.append(series)
                location += sum(data_source.lengths[i][j])

            worksheet.add_chart(chart, 'D8')


def _select_processing_options(data_sources):
    """
    Launches a window to select the processing options.

    Parameters
    ----------
    data_sources : list or tuple
        A container (list, tuple) of DataSource objects.

    Returns
    -------
    values : dict
        A dictionary containing the processing options.

    """

    if _get_save_location().joinpath('previous_search.json').exists():
        last_search_disabled = False
    else:
        last_search_disabled = True

    #Selection of check boxes
    options_layout = [
        [sg.Text('Select Input', relief='ridge', justification='center',
                 size=(40, 1))],
        [sg.Radio('Multiple Files', 'options_radio', default=True,
                  key='multiple_files', enable_events=True)],
        [sg.Check('Use Previous Search', key='use_last_search',
                  disabled=last_search_disabled, pad=((40, 0), (1, 0)))],
        [sg.Radio('Single File', 'options_radio', key='single_file',
                  enable_events=True)],
        [sg.Text('Select All Boxes That Apply', relief='ridge',
                 justification='center', size=(40, 1))],
        [sg.Check('Process Data', key='process_data', default=True,
                  enable_events=True)],
        [sg.Check('Fit Peaks', key='fit_peaks', enable_events=True)],
        [sg.Check('Save to Excel', key='save_fitting', pad=((40, 0), (1, 0)),
                  enable_events=True, disabled=True)],
        [sg.Check('Plot in Python', key='plot_python')],
        [sg.Check('Move File(s)', key='move_files', default=False)],
        [sg.Check('Save Excel File', key='save_excel',
                  default=True, enable_events=True),
         sg.Combo(('Create new file', 'Append to existing file'),
                  key='append_file', readonly=True,
                  default_value='Append to existing file', size=(19, 1))],
        [sg.Check('Plot Data in Excel', key='plot_data_excel',
                  pad=((40, 0), (1, 0)))],
        [sg.Check('Plot Fit Results in Excel', key='plot_fit_excel',
                  disabled=True, pad=((40, 0), (1, 0)))],
        [sg.Input('', key='file_name', visible=False),
         sg.Input('', key='save_as', visible=False,
                  enable_events=True, do_not_clear=False),
         sg.SaveAs(file_types=(("Excel Workbook (xlsx)", "*.xlsx"),),
                   key='file_save_as', target='save_as',
                   pad=((40, 0), 5))],
    ]

    data_sources_radios = [
        [sg.Radio(f'{j + 1}) {source.name}', 'radio', key=f'source_{source.name}',
                  enable_events=True)] for j, source in enumerate(data_sources)
    ]

    layout = [
        [sg.TabGroup([
            [sg.Tab('Options', options_layout, key='tab1'),
             sg.Tab('Data Sources', [
                [sg.Text('Select Data Source:', relief='ridge',
                         justification='center', size=(40, 1))],
                *data_sources_radios,
             ], key='tab2')]
        ], tab_background_color=sg.theme_background_color(), key='tab')],
        [sg.Button('Next', bind_return_key=True,
                   button_color=utils.PROCEED_COLOR)]
    ]

    window = sg.Window('Main Menu', layout)

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED:
            utils.safely_close_window(window)

        elif event == 'Next':
            if any((values['fit_peaks'], values['plot_python'],
                    values['save_excel'], values['move_files'],
                    values['process_data'])):

                close_window = False
                for source in data_sources:
                    if values[f'source_{source.name}']:
                        close_window = True
                        break
                if close_window:
                    if not values['save_excel'] or values['file_name']:
                        break
                    else:
                        sg.popup(
                            'Please select a filename for the output Excel file.\n',
                            title='Error'
                        )
                else:
                    sg.popup('Please select a data source.\n',
                             title='Error')

            elif values['move_files']:
                break

            else:
                sg.popup('Please select a data processing option.\n',
                         title='Error')

        if event == 'multiple_files':
            if not last_search_disabled:
                window['use_last_search'].update(disabled=False)

        elif event == 'single_file':
            window['use_last_search'].update(value=False, disabled=True)

        elif event == 'fit_peaks':
            if values['fit_peaks'] and values['save_excel']:
                window['save_fitting'].update(value=True, disabled=False)
                window['plot_fit_excel'].update(disabled=False)
            else:
                window['save_fitting'].update(value=False, disabled=True)
                window['plot_fit_excel'].update(value=False, disabled=True)

        elif event == 'save_fitting':
            if values['save_fitting']:
                window['plot_fit_excel'].update(disabled=False)
            else:
                window['plot_fit_excel'].update(value=False, disabled=True)

        elif event == 'save_excel':
            if values['save_excel']:
                window['append_file'].update(readonly=True)
                window['plot_data_excel'].update(disabled=False)

                if values['fit_peaks']:
                    window['save_fitting'].update(value=True, disabled=False)
                    window['plot_fit_excel'].update(disabled=False)
            else:
                window['append_file'].update(
                    value='Append to existing file', disabled=True
                )
                window['plot_data_excel'].update(value=False, disabled=True)
                window['plot_fit_excel'].update(value=False, disabled=True)
                window['save_fitting'].update(value=False, disabled=True)

        elif event == 'save_as' and values['save_as']:
            file_path = Path(values['save_as'])
            if file_path.suffix.lower() != '.xlsx':
                file_path = Path(file_path.parent, file_path.stem + '.xlsx')
            window['file_name'].update(value=str(file_path))

    window.close()
    del window

    values['append_file'] = values['append_file'] == 'Append to existing file'

    # removes unneeded keys
    for key in ('file_save_as', 'save_as', 'single_file', 'tab'):
        del values[key]

    return values


def _create_column_labels_window(dataset, data_source, options, index,
                                 gui_inputs, location, last_index):
    """
    Creates the window to specify the sample and column labels.

    Parameters
    ----------
    dataset : list
        The list of lists of dataframes for one dataset.
    data_source : DataSource
        The DataSource object for the data.
    options : dict
        The dictionary that contains information about which
        processing steps will be conducted.
    index : int
        The index of the dataset within the total list of datasets.
    gui_inputs : dict
        A dictionary of values to overwrite the default gui values, used
        when displaying a previous window.
    location : tuple
        The window location.
    last_index : bool
        If True, designates that it is the last index.

    Returns
    -------
    validations : dict
        A dictionary with the validations needed for the created window.
    sg.Window
        The created window to select the labels.

    """

    labels = data_source.create_needed_labels(
        max(len(df.columns) for sample in dataset for df in sample)
    )

    available_cols = labels[0] + labels[1] if options['process_data'] else labels[0]
    if (data_source.x_plot_index >= len(available_cols)
            or data_source.y_plot_index >= len(available_cols)):
        x_plot_index = 0
        y_plot_index = len(available_cols) - 1
    else:
        x_plot_index = data_source.x_plot_index
        y_plot_index = data_source.y_plot_index

    validations = {'user_inputs': []}
    default_inputs = {
        'x_plot_index': x_plot_index,
        'y_plot_index': y_plot_index,
        'x_min': '',
        'x_max': '',
        'y_min': '',
        'y_max': '',
        'x_label': available_cols[x_plot_index],
        'y_label': available_cols[y_plot_index],
        'x_log_scale': False,
        'y_log_scale': False,
        'chart_title': ''
    }

    for i in range(len(dataset)):
        default_inputs.update({f'sample_name_{i}': ''})
        validations['user_inputs'].append([
            f'sample_name_{i}', f'sample name {i + 1}', utils.string_to_unicode, True, None
        ])

    if options['process_data'] and data_source.dataset_summary_functions:
        default_inputs.update({'summary_name': 'Summary'})
        validations['user_inputs'].append([
            'summary_name', 'summary name',
            utils.string_to_unicode, True, None
        ])

    keys = ('data_label', 'calculation_label',
            'sample_summary_label', 'dataset_summary_label')

    for i, label_list in enumerate(labels):
        default_inputs.update(
            {f'{keys[i]}_{j}': label for j, label in enumerate(label_list)}
        )

    default_inputs.update(gui_inputs)
    if 'sheet_name' not in default_inputs:
        default_inputs['sheet_name'] = f'Sheet {index + 1}'

    if options['save_excel']:
        header = 'Sheet Name: '
        header_visible = True
        validations['user_inputs'].append([
            'sheet_name', 'sheet name', utils.string_to_unicode, False, None
        ])
    else:
        header = f'Dataset {index + 1}'
        header_visible = False

    column_width = 32
    labels_layout = [
        [sg.Text(header, visible=header_visible),
         sg.Input(default_inputs['sheet_name'], key='sheet_name',
                  size=(15, 1), visible=header_visible)],
        [sg.Text('Sample Names', size=(column_width, 1),
                 justification='center', relief='ridge', pad=(5, 10))]
    ]

    for i in range(len(dataset)):
        labels_layout.append(
            [sg.Text(f'    Sample {i + 1}'),
             sg.Input(default_inputs[f'sample_name_{i}'], size=(20, 1),
                      key=f'sample_name_{i}')]
        )

    if options['process_data'] and data_source.dataset_summary_functions:
        labels_layout.append(
            [sg.Text('    Summary'),
             sg.Input(default_inputs['summary_name'], size=(20, 1),
                      key='summary_name')]
        )

    labels_layout.extend([
        [sg.Text('Column Labels', size=(column_width, 1),
                 justification='center', relief='ridge', pad=(5, 10))],
        [sg.Text('Imported Data Labels:')]
    ])

    for i in range(len(labels[0])):
        validations['user_inputs'].append([
            f'{keys[0]}_{i}', f'raw data label {i}', utils.string_to_unicode, True, None
        ])
        labels_layout.append(
            [sg.Text(f'    Column {i}'),
             sg.Input(default_inputs[f'{keys[0]}_{i}'], size=(20, 1),
                      key=f'{keys[0]}_{i}')]
        )

    if options['process_data']:
        calc_labels = ('Calculation', 'Sample Summary', 'Dataset Summary')

        for j, label_list in enumerate(labels[1:]):
            if label_list:
                labels_layout.append([sg.Text(f'{calc_labels[j]} Labels:')])
                for k in range(len(label_list)):
                    col_num = i + 1 + k if j == 0 else k # continue column numbering for Calculations
                    labels_layout.append(
                        [sg.Text(f'    Column {col_num}'),
                         sg.Input(default_inputs[f'{keys[j + 1]}_{k}'], size=(20, 1),
                                  key=f'{keys[j + 1]}_{k}')]
                    )
                    validations['user_inputs'].append([
                        f'{keys[j + 1]}_{k}', f'{calc_labels[j].lower()} label {col_num}',
                        utils.string_to_unicode, True, None
                    ])

    labels_column = [
        sg.Column(labels_layout, scrollable=True,
                  vertical_scroll_only=True, size=(404, 400))
    ]

    if not options['plot_data_excel']:
        main_section = [sg.Frame('', [labels_column])]
    else:
        validations['integers'] = [
            ['x_plot_index', 'x plot index'],
            ['y_plot_index', 'y plot index']
        ]
        validations['user_inputs'].extend([
            ['x_min', 'x min', float , True, None],
            ['x_max', 'x max', float , True, None],
            ['y_min', 'y min', float , True, None],
            ['y_max', 'y max', float , True, None],
            ['x_label', 'x axis label', utils.string_to_unicode, False, None],
            ['y_label', 'y axis label', utils.string_to_unicode, False, None],
            ['chart_title', 'chart title', utils.string_to_unicode, True, None]
        ])

        plot_layout = [
            [sg.Text('Chart title:'),
             sg.Input(default_inputs['chart_title'], key='chart_title', size=(20, 1))],
            [sg.Text('Column of x data for plotting:'),
             sg.Combo(list(range(len(available_cols))),
                      key='x_plot_index', readonly=True, size=(3, 1),
                      default_value=default_inputs['x_plot_index'])],
            [sg.Text('Column of y data for plotting:'),
             sg.Combo(list(range(len(available_cols))),
                      key='y_plot_index', readonly=True, size=(3, 1),
                      default_value=default_inputs['y_plot_index'])],
            [sg.Text('X axis label:'),
             sg.Input(default_inputs['x_label'], key='x_label', size=(20, 1))],
            [sg.Text('Y axis label:'),
             sg.Input(default_inputs['y_label'], key='y_label', size=(20, 1))],
            [sg.Text(("Min and max values to show on the plot\n"
                      "(leave blank to use Excel's default):"))],
            [sg.Text('    X min:', size=(8, 1)),
             sg.Input(default_inputs['x_min'], key='x_min', size=(5, 1)),
             sg.Text('    X max:', size=(8, 1)),
             sg.Input(default_inputs['x_max'], key='x_max', size=(5, 1))],
            [sg.Text('    Y min:', size=(8, 1)),
             sg.Input(default_inputs['y_min'], key='y_min', size=(5, 1)),
             sg.Text('    Y max:', size=(8, 1)),
             sg.Input(default_inputs['y_max'], key='y_max', size=(5, 1))],
            [sg.Text('Use logorithmic scale?')],
            [sg.Check('X axis', default_inputs['x_log_scale'],
                      key='x_log_scale', pad=((20, 5), 5)),
             sg.Check('Y axis', default_inputs['y_log_scale'], key='y_log_scale')]
        ]

        main_section = [
            sg.TabGroup([
                [sg.Tab('Labels', [labels_column], key='tab1'),
                 sg.Tab('Excel Plot', plot_layout, key='tab2')]
            ], key='tab', tab_background_color=sg.theme_background_color())
        ]

    layout = [
        [sg.Menu([['&Help', ['&Unicode Help']]], key='-menu-',
                 background_color=sg.theme_background_color())],
        main_section,
        [sg.Text('')],
        [sg.Button('Back', disabled=index == 0),
         sg.Button('Finish' if last_index else 'Next', bind_return_key=True,
                   button_color=utils.PROCEED_COLOR)]
    ]

    return validations, sg.Window(f'Dataset {index + 1} Options',
                                  layout, location=location)


def _select_column_labels(dataframes, data_source, processing_options):
    """
    Handles the event loop for the window to select the sample and column labels.

    Parameters
    ----------
    dataframes : list
        A list of lists of lists of pd.DataFrame objects, containing the all
        the data to process.
    data_source : DataSource
        The DataSource object for the data.
    processing_options : dict
        The dictionary that contains information about which
        processing steps will be conducted.

    Returns
    -------
    label_values : list
        A list of dictionaries containing all of the sample and column
        labels, as well as the Excel plot options, if plotting in Excel.
        Each entry in the list corresponds to one dataset.

    """

    label_values = [{} for _ in dataframes]
    location = (None, None)
    for i, dataset in enumerate(dataframes):
        j = i

        validations, window = _create_column_labels_window(
            dataset, data_source, processing_options, i, label_values[i],
            location, i==len(dataframes) - 1
        )

        while True:
            event, values = window.read()

            if event == sg.WIN_CLOSED:
                utils.safely_close_window(window)

            elif event == 'Unicode Help':
                sg.popup(
                    ('"\\u00B2": \u00B2 \n"\\u03B8": \u03B8 \n"'
                     '\\u00B0": \u00B0\n"\\u03bc": \u03bc\n"\\u03bb": \u03bb\n'
                     '\nFor example, Acceleration'
                     ' (m/s\\u00B2) creates Acceleration (m/s\u00B2).\n'),
                    title='Example Unicode'
                )

            elif event in ('Back', 'Next', 'Finish'):
                if utils.validate_inputs(values, **validations):
                    label_values[j].update(values)
                    location = window.current_location()
                    window.close()

                    if event == 'Back':
                        j -= 1
                    else:
                        j += 1

                    if j <= i:
                        validations, window = _create_column_labels_window(
                            dataframes[j], data_source, processing_options, j,
                            label_values[j], location, j==len(dataframes) - 1
                        )
                    else:
                        if i < len(dataframes) - 1:
                            label_values[i + 1].update(values)
                            label_values[i + 1].pop('sheet_name')
                        break

        window.close()
        window = None

    return label_values


def _collect_column_labels(dataframes, data_source, labels, options):
    """
    Collects all labels and condenses them into a single list of labels per dataset.

    Also adds in blank labels for spacer columns between entries and samples.

    Parameters
    ----------
    dataframes : list
        A list of lists of lists of pd.DataFrame objects, containing the all
        the data to process.
    data_source : DataSource
        The DataSource object for the data.
    labels : list(dict)
        A list of dictionaries. Each dictionary contains all of the
        sample names and column labels for a dataset.
    options : dict
        The dictionary that contains information about which
        processing steps will be conducted.

    """

    for i, dataset in enumerate(dataframes):
        labels[i]['total_labels'] = []

        for j in range(len(labels[i]['sample_names'])):
            index_modifier = -1 if labels[i]['sample_summary_labels'] else 0

            for entry_num in range(1, len(dataset[j]) + 1 + index_modifier):
                for label in labels[i]['column_labels']:
                    if data_source.label_entries and len(dataset[j]) > 1 and label:
                        labels[i]['total_labels'].append(f'{label} {entry_num}')
                    else:
                        labels[i]['total_labels'].append(label)

                if options['process_data'] and entry_num != len(dataset[j]) + index_modifier:
                    labels[i]['total_labels'].extend([
                        '' for _ in range(data_source.entry_separation)
                    ])

            if options['process_data']:
                if labels[i]['sample_summary_labels']:
                    labels[i]['total_labels'].extend([
                        *['' for _ in range(data_source.entry_separation)],
                        *labels[i]['sample_summary_labels']
                    ])

                labels[i]['total_labels'].extend([
                    '' for _ in range(data_source.sample_separation)
                ])

        if options['process_data'] and labels[i]['dataset_summary_labels']:
            labels[i]['total_labels'].extend([
                *labels[i]['dataset_summary_labels'],
                *['' for _ in range(data_source.sample_separation)]
            ])


def _fit_data(datasets, data_source, labels, excel_writer, options):
    """
    Handles peak fitting and any exceptions that occur during peak fitting.

    Parameters
    ----------
    dataframes : list
        A nested list of lists of lists of dataframes.
    data_source : DataSource
        The selected DataSource.
    labels : list(dict)
        A list of dictionaries containing the sample names and column
        labels for each dataset.
    excel_writer : pd.ExcelWriter
        The pandas ExcelWriter object that contains all of the
        information about the Excel file being created.
    options : dict
        A dictionary containing the relevent keys 'save_fitting' and
        'plot_fit_excel' which determine whether the fit results
        will be saved to Excel and whether the results will be plotted,
        respectively.

    Returns
    -------
    results : list
        A nested list of lists of lists, one entry for each entry in each sample
        in each dataset in datasets. If fitting was not done for the entry,
        the value will be None.

    Raises
    ------
    utils.WindowCloseError
        Raised if fitting was ended early by the user.

    """

    # Changes some defaults for the plot formatting to look nice.
    mpl_changes = { #TODO maybe allow this to be an input into the main gui function
        'font.serif': 'Times New Roman',
        'font.family': 'serif',
        'font.size': 12,
        'mathtext.default': 'regular',
        'xtick.direction': 'in',
        'ytick.direction': 'in',
        'xtick.minor.visible': True,
        'ytick.minor.visible': True,
        'xtick.major.size': 5,
        'xtick.major.width': 0.6,
        'xtick.minor.size': 2.5,
        'xtick.minor.width': 0.6,
        'ytick.major.size': 5,
        'ytick.major.width': 0.6,
        'ytick.minor.size': 2.5,
        'ytick.minor.width': 0.6,
        'lines.linewidth': 2,
        'lines.markersize': 5,
        'axes.linewidth': 0.6,
        'legend.frameon': False,
        'figure.dpi': 150,
        'figure.figsize': (6, 4.5)
    }

    results = [[[] for sample in dataset] for dataset in datasets]

    # Allows exiting from the peak fitting GUI early, if desired or because of
    # an exception, while still continuing with the program.
    try:
        default_inputs = {
            'x_fit_index': data_source.x_plot_index,
            'y_fit_index': data_source.y_plot_index
        }

        for i, dataset in enumerate(datasets):
            default_inputs.update({
                'x_label': labels[i]['column_labels'][data_source.x_plot_index],
                'y_label': labels[i]['column_labels'][data_source.y_plot_index]
            })

            for j, sample in enumerate(dataset):
                sample_names = labels[i]['sample_names'] + labels[i]['summary_name']
                for k, entry in enumerate(sample):
                    if len(sample) > 1:
                        name = f'{sample_names[j]}_{k + 1}_fit'
                    else:
                        name = sample_names[j]

                    default_inputs.update({'sample_name': name})

                    fit_output, default_inputs, proceed = launch_peak_fitting_gui(
                        entry, default_inputs, excel_writer,
                        options['save_fitting'], options['plot_fit_excel'],
                        mpl_changes, False
                    )

                    results[i][j].extend(fit_output)

                    if not proceed:
                        raise utils.WindowCloseError

    except (utils.WindowCloseError, KeyboardInterrupt):
        print('\nPeak fitting manually ended early.\nMoving on with program.')

    except Exception:
        print('\nException occured during peak fitting:\n')
        print(traceback.format_exc())
        print('Moving on with program.')

    return results


def _plot_data(datasets, data_source):
    """
    Handles plotting and any exceptions that occur during plotting.

    Parameters
    ----------
    datasets : list
        A nested list of lists of lists of dataframes.
    data_source : DataSource
        The DataSource object whose figure_rcParams attribute will be used
        to set matplotlib's rcParams.

    Returns
    -------
    list
        A nested list of lists, with one entry per dataset in datasets.
        Each entry contains the matplotlib Figure, and a dictionary
        containing the Axes. If plotting was exited before plotting all
        datasets in dataframes, then [None, None] will be the entry instead.

    """

    plot_datasets = []
    for dataset in datasets: # Flattens the dataset to a single list per dataset
        plot_datasets.append(list(itertools.chain.from_iterable(dataset)))

    return launch_plotting_gui(plot_datasets, data_source.figure_rcParams)


def _move_files(files):
    """
    Launches a window to select the new folder destinations for the files.

    Parameters
    ----------
    files : list
        A nested list of lists of lists of strings corresponding
        to file paths.

    """

    text_layout = [[sg.Text(f'Dataset {i + 1}')] for i in range(len(files))]
    files_layout = [
        [sg.Input('', key=f'folder_{i}', enable_events=True,
                  disabled=True),
         sg.FolderBrowse(target=f'folder_{i}', key=f'button_{i}')]
        for i in range(len(files))
    ]
    tot_layout = [i for j in zip(text_layout, files_layout) for i in j]

    if len(files) > 2:
        scrollable = True
        size = (600, 200)
    else:
        scrollable = False
        size = (None, None)

    layout = [
        [sg.Text('Choose the folder(s) to move files to:', size=(30, 1))],
        [sg.Frame('', [[sg.Column(tot_layout, scrollable=scrollable,
                                  vertical_scroll_only=True, size=size)]])],
        [sg.Button('Submit', bind_return_key=True,
                   button_color=utils.PROCEED_COLOR),
         sg.Check('All Same Folder', key='same_folder',
                  enable_events=True, disabled=len(files) == 1)]
    ]

    try:
        window = sg.Window('Move Files', layout)
        while True:
            event, values = window.read()

            if event == sg.WIN_CLOSED:
                utils.safely_close_window(window)

            elif event.startswith('folder_') and values['same_folder']:
                for i in range(1, len(files)):
                    window[f'folder_{i}'].update(value=values['folder_0'])

            elif event == 'same_folder':
                if values['same_folder']:
                    for i in range(1, len(files)):
                        window[f'folder_{i}'].update(value=values['folder_0'])
                        window[f'button_{i}'].update(disabled=True)
                else:
                    for i in range(1, len(files)):
                        window[f'button_{i}'].update(disabled=False)

            elif event == 'Submit':
                if any(not values[key] for key in values if key.startswith('folder_')):
                    sg.popup('Please enter folders for all datasets', title='Error')
                else:
                    break

        window.close()
        del window

    except (utils.WindowCloseError, KeyboardInterrupt):
        print('\nMoving files manually ended early.\nMoving on with program.')

    else:
        try:
            folders = [values[f'folder_{i}'] for i in range(len(files))]
            for i, file_list in enumerate(files):
                # Will automatically rename files if there is already a file with
                # the same name in the destination folder.
                file_mover(file_list, new_folder=folders[i], skip_same_files=False)
        except Exception:
            print('\nException occured during moving files:\n')
            print(traceback.format_exc())
            print('Moving on with program.')


def launch_main_gui(data_sources):
    """
    Goes through all of the windows to find files, process/plot/fit data, and save to Excel.

    Parameters
    ----------
    data_sources : list(DataSource) or tuple(DataSource)
        A container (list, tuple) of mcetl.DataSource objects.

    Returns
    -------
    output : dict
        A dictionary containing the following keys and values:
            dataframes : list or None
                A list of lists of dataframes, with each dataframe containing the
                data imported from a raw data file; will be None if the function
                fails before importing data, or if the only processing step taken
                was moving files.
            fit_results : list or None
                A nested list of lists of lmfit ModelResult objects, with each
                ModelResult pertaining to a single fitting, each list of
                ModelResults containing all of the fits for a single dataset,
                and east list of lists pertaining the data within one processed
                dataframe; will be None if fitting is not done,  or only
                partially filled if the fitting process ends early.
            plot_results : list or None
                A list of lists, with one entry per dataset. Each interior
                list is composed of a matplotlib.Figure object and a
                dictionary of matplotlib.Axes objects. Will be None if
                plotting is not done, or only partially filled if the plotting
                process ends early.
            writer : pd.ExcelWriter or None
                The pandas ExcelWriter used to create the output Excel file; will
                be None if the output results were not saved to Excel.

    Notes
    -----
    The entire function is wrapped in a try-except block. If the user exits the
    program early by exiting out of a GUI, a custom WindowCloseError exception is
    thrown, which is just passed, allowing the program is close without error.
    If other exceptions occur, their traceback is printed.

    """

    output = {
        'dataframes': None,
        'fit_results': None,
        'plot_results': None,
        'writer': None
    }

    if not isinstance(data_sources, (list, tuple)):
        data_sources = [data_sources]
    if any(not isinstance(data_source, DataSource) for data_source in data_sources):
        raise TypeError("Only DataSource objects can be used in the main gui.")

    try:
        processing_options = _select_processing_options(data_sources)

        # Specifying the selected data source
        for source in data_sources:
            if processing_options[f'source_{source.name}']:
                data_source = source
                break

        # Selection of raw data files
        if processing_options['multiple_files']:
            if processing_options['use_last_search']:
                with _get_save_location().joinpath('previous_search.json').open('r') as f:
                    files = json.load(f)
            else:
                files = file_finder(
                    file_type=data_source.file_type, num_files=data_source.num_files
                )

                # Saves the last search to a json file so it can be used again to bypass the search.
                save_path = _get_save_location()
                save_path.mkdir(exist_ok=True)
                with save_path.joinpath('previous_search.json').open('w') as f:
                    json.dump(files, f, indent=2)

            # Imports the raw data from the files
            if any((processing_options['process_data'],
                    processing_options['save_excel'],
                    processing_options['fit_peaks'],
                    processing_options['plot_python'])):

                output['dataframes'] = [[[] for sample in dataset] for dataset in files]
                import_vals = [[[] for sample in dataset] for dataset in files]
                if files[0][0][0].endswith('.xlsx'):
                    for i, dataset in enumerate(files):
                        for j, sample in enumerate(dataset):
                            for entry in sample:
                                #disable_blank_col = not (i == 0 and j == 0) #TODO use this later to lock out changing the number of columns
                                import_values = utils.select_file_gui(
                                    data_source, sample
                                )
                                added_dataframes = utils.raw_data_import(
                                    import_values, sample, False
                                )
                                output['dataframes'][i][j].extend(added_dataframes)
                                import_vals[i][j].extend(
                                    [import_values] * len(added_dataframes)
                                )

                else:
                    import_values = utils.select_file_gui(data_source, files[0][0][0])
                    for i, dataset in enumerate(files):
                        for j, sample in enumerate(dataset):
                            for entry in sample:
                                output['dataframes'][i][j].extend(
                                    utils.raw_data_import(import_values, entry, False)
                                )
                                import_vals[i][j].append(import_values)

        else:
            import_values = utils.select_file_gui(data_source)
            output['dataframes'] = [[
                utils.raw_data_import(import_values, import_values['file'], False)
            ]]
            files = [[[import_values['file']]]]
            import_vals = [[[import_values] * len(output['dataframes'][0][0])]]

        # Specifies column names
        if any((processing_options['process_data'],
                processing_options['save_excel'],
                processing_options['fit_peaks'],
                processing_options['plot_python'])):

            label_values = _select_column_labels(
                output['dataframes'], data_source, processing_options
            )

            labels = [{} for _ in output['dataframes']]
            plot_options = []
            for i, values in enumerate(label_values):
                labels[i]['sheet_name'] = values['sheet_name']
                labels[i]['sample_names'] = [
                    values[key] for key in values if key.startswith('sample_name')
                ]
                labels[i]['column_labels'] = [
                    *[values[key] for key in values if key.startswith('data_label')],
                    *[values[key] for key in values if key.startswith('calculation_label')]
                ]
                labels[i]['sample_summary_labels'] = [
                    values[key] for key in values if key.startswith('sample_summary_label')
                ]
                labels[i]['dataset_summary_labels'] = [
                    values[key] for key in values if key.startswith('dataset_summary_label')
                ]
                labels[i]['summary_name'] = [
                    values[key] for key in values if key == 'summary_name'
                ]

                if not processing_options['plot_data_excel']:
                    plot_options.append(None)
                else:
                    plot_options.append({
                        'x_label': values['x_label'],
                        'y_label': values['y_label'],
                        'chart_title' : values['chart_title'],
                        'x_plot_index': values['x_plot_index'],
                        'y_plot_index': values['y_plot_index'],
                        'x_min': values['x_min'] if values['x_min'] != '' else None,
                        'x_max': values['x_max'] if values['x_max'] != '' else None,
                        'y_min': values['y_min'] if values['y_min'] != '' else None,
                        'y_max': values['y_max'] if values['y_max'] != '' else None,
                        'x_log_scale': values['x_log_scale'],
                        'y_log_scale': values['y_log_scale']
                    })

            if not processing_options['process_data']: # Otherwise, will assign labels after Separation functions
                _collect_column_labels(output['dataframes'], data_source,
                                       labels, processing_options)

        if processing_options['save_excel'] or processing_options['process_data']:

            if processing_options['process_data']:
                # Perform separation functions
                output['dataframes'], import_vals = data_source.do_separation_functions(
                    output['dataframes'], import_vals
                )
                # Assign reference indices for all relevant columns
                data_source.set_references(output['dataframes'], import_vals)

                _collect_column_labels(output['dataframes'], data_source, labels, processing_options)

            # Merge dataframes for each dataset
            merged_dataframes = data_source.merge_datasets(output['dataframes'])
            output['dataframes'] = None # Frees up memory

            if processing_options['save_excel'] and processing_options['process_data']:
                merged_dataframes = data_source.do_excel_functions(merged_dataframes)

            if processing_options['save_excel']:
                if (processing_options['append_file']
                        and Path(processing_options['file_name']).exists()):
                    mode = 'a'
                else:
                    processing_options['append_file'] = False
                    mode = 'w'

                output['writer'] =  pd.ExcelWriter(
                    processing_options['file_name'], engine='openpyxl', mode=mode
                )
                # Formatting styles for the Excel workbook
                for style, kwargs in data_source.excel_formats.items():
                    try:
                        output['writer'].book.add_named_style(NamedStyle(style, **kwargs))
                    except ValueError: # Style already exists in the workbook
                        pass

                _write_to_excel(
                    merged_dataframes, data_source, labels, output['writer'],
                    processing_options['plot_data_excel'], plot_options
                )

            if processing_options['process_data']:
                merged_dataframes = data_source.do_python_functions(merged_dataframes)

            # Split data back into individual dataframes
            output['dataframes'] = data_source.split_into_entries(merged_dataframes)
            del merged_dataframes

        # Assigns column names to the dataframes
        if any((processing_options['process_data'],
                processing_options['fit_peaks'],
                processing_options['plot_python'])):

            pass #TODO later assign column headers for all dfs based on labels['total_columns']

        """
        #renames dataframe columns if there are repeated terms,
        #since it causes issues for plotting and fitting
        if any((plot_python, fit_peaks)):

            for k, dataframe in enumerate(dataframes):
                header_list = dataframe.columns.tolist()
                for i, header in enumerate(header_list):
                    num = 1
                    if header in header_list[i+1:]:
                        for j, elem in enumerate(header_list[i+1:]):
                            if header == elem:
                                header_list[i+1+j] = f'{header}_{num}'
                                num += 1
                dataframes[k].columns = header_list
        """

        # Handles peak fitting
        if processing_options['fit_peaks']:
            output['fit_results'] = _fit_data(
                output['dataframes'], data_source, labels, output['writer'], processing_options
            )

        # Handles saving the Excel file
        if processing_options['save_excel']:
            utils.save_excel_file(output['writer'])

        # Handles moving files
        if processing_options['move_files']:
            _move_files(files)

        # Handles plotting in python
        if processing_options['plot_python']:
            output['plot_results'] = _plot_data(output['dataframes'], data_source) #TODO later pass labels

    except (utils.WindowCloseError, KeyboardInterrupt):
        pass
    except Exception:
        print(traceback.format_exc())

    return output
