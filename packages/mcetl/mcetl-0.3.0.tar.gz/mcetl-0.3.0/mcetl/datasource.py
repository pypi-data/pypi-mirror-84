# -*- coding: utf-8 -*-
"""Contains the DataSource class.

@author: Donald Erb
Created on Jul 31, 2020

#TODO need to update all docstrings
"""


import itertools

import numpy as np
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd

from . import utils
from .functions import SeparationFunction, CalculationFunction, SummaryFunction


class DataSource:
    """
    Used to give default settings for importing data and various functions based on the source.

    Attributes
    ----------
    lengths : list
        A list of lists of lists of integers, corresponding to the number of columns
        in each individual entry in the total dataframes for the DataSource.
        Used to split the concatted dataframe back into individual dataframes for
        each dataset.
    references : list
        A list of dictionaries, with each dictionary containing the column numbers
        for each unique variable and calculation for the merged dataframe of each
        dataset.

    """

    def __init__(
            self,
            name,
            functions=None,
            column_labels=None,
            column_numbers=None,
            start_row=0,
            end_row=0,
            separator=None,
            file_type=None,
            num_files=1,
            unique_variables=None,
            unique_variable_indices=None,
            xy_plot_indices=None,
            figure_rcParams=None,
            excel_writer_formats=None,
            excel_row_offset=0,
            excel_column_offset=0,
            entry_separation=0,
            sample_separation=0,
            label_entries=True):
        """
        DataSource initialization.

        Parameters
        ----------
        name : TYPE
            DESCRIPTION.
        column_labels : TYPE, optional
            DESCRIPTION. The default is None.
        functions : TYPE, optional
            DESCRIPTION. The default is None.
        column_numbers : TYPE, optional
            DESCRIPTION. The default is None.
        start_row : TYPE, optional
            DESCRIPTION. The default is 0.
        end_row : TYPE, optional
            DESCRIPTION. The default is 0.
        separator : TYPE, optional
            DESCRIPTION. The default is None.
        unique_variable_indices : TYPE, optional
            DESCRIPTION. The default is None.
        xy_plot_indices : TYPE, optional
            DESCRIPTION. The default is None.
        file_type : TYPE, optional
            DESCRIPTION. The default is None.
        num_files : TYPE, optional
            DESCRIPTION. The default is 1.
        unique_variables : TYPE, optional
            DESCRIPTION. The default is None.
        figure_rcParams : dict, optional
            A dictionary containing any changes to matplotlib's rcParams.
            The default is None.
        excel_writer_formats : dict(dict), optional
            A dictionary of styles to overwrite the default settings for the
            pd.ExcelWriter. The following keys are used when writing data
            from files to Excel:
                'header_even', 'header_odd', 'subheader_even', 'subheader_odd',
                'columns_even', 'columns_odd'
            The following keys are used when writing data from mcetl's peak
            fitting to Excel:
                'fitting_header_even', 'fitting_header_odd', 'fitting_subheader_even',
                'fitting_subheader_odd', 'fitting_columns_even', 'fitting_columns_odd',
                'fitting_descriptors_even', 'fitting_descriptors_odd'
            The values for the dictionaries must also be dictionaries, with
            keys corresponding to keyword inputs for openpyxl's NamedStyle.
        sample_separation : TYPE, optional
            DESCRIPTION. The default is 0.
        entry_separation : TYPE, optional
            DESCRIPTION. The default is 0.
        label_entries : bool, optional
            If True, will add a number to the column labels for each
            entry in a sample.

        Raises
        ------
        ValueError
            Raised if the input name is a blank string, or if either excel_row_offset
            or excel_column_offset is < 0.
        TypeError
            DESCRIPTION.
        IndexError
            DESCRIPTION.

        """

        if name:
            self.name = name
        else:
            raise ValueError('DataSource name cannot be a blank string.')

        # attributes that will be set later
        self.lengths = None
        self.excel_formats = None
        self.references = None

        self.start_row = start_row
        self.end_row = end_row
        self.file_type = file_type
        self.num_files = num_files
        self.sample_separation = sample_separation
        self.entry_separation = entry_separation
        self.label_entries = label_entries
        self.column_labels = column_labels if column_labels is not None else []
        self.figure_rcParams = figure_rcParams if figure_rcParams is not None else {}

        # Turns the separator into a raw string equivalent so that it properly displays in the GUIs
        if separator is not None:
            for replacement in (('\\', '\\\\'), ('\n', '\\n'), ('\t', '\\t'), ('\r', '\\r')):
                separator = separator.replace(*replacement)
        self.separator = separator

        # Ensures excel_row_offset and excel_column_offset are >= 0
        if any(value < 0 for value in (excel_column_offset, excel_row_offset)):
            raise ValueError('excel_column_offset and excel_row_offset must be >= 0.')
        else:
            self.excel_row_offset = excel_row_offset
            self.excel_column_offset = excel_column_offset

        if unique_variables is None:
            self.unique_variables = []
        elif isinstance(unique_variables, str):
            self.unique_variables = [unique_variables]
        else:
            self.unique_variables = unique_variables

        if column_numbers is not None:
            self.column_numbers = column_numbers
        else:
            self.column_numbers = list(range(len(self.unique_variables)))

        # ensures the number of imported columns can accomodate all variables
        if len(self.column_numbers) < len(self.unique_variables):
            raise IndexError((
                f'The number of columns specified for mcetl.DataSource "{self.name}" must '
                'be greater or equal to the number of unique variables, not less than.'
            ))

        # sorts the functions by their usage
        self.separation_functions = []
        self.calculation_functions = []
        self.sample_summary_functions = []
        self.dataset_summary_functions = []
        if functions is not None:
            if not isinstance(functions, (list, tuple)):
                functions = (functions,)
            for function in functions:
                if isinstance(function, SummaryFunction):
                    if function.sample_summary:
                        self.sample_summary_functions.append(function)
                    else:
                        self.dataset_summary_functions.append(function)
                elif isinstance(function, CalculationFunction):
                    self.calculation_functions.append(function)
                elif isinstance(function, SeparationFunction):
                    self.separation_functions.append(function)
                else:
                    raise TypeError(f'{function} is not a valid Function object.')

        self._validate_target_columns()

        # indices for each unique variable for data processing
        if unique_variable_indices is None:
            self.unique_variable_indices = list(range(len(self.unique_variables)))
        elif isinstance(unique_variable_indices, (str, int)):
            self.unique_variable_indices = [unique_variable_indices]
        else:
            self.unique_variable_indices = unique_variable_indices

        while len(self.unique_variables) > len(self.unique_variable_indices):
            self.unique_variable_indices.append(max(self.unique_variable_indices) + 1)

        # x and y indices for plotting
        if isinstance(xy_plot_indices, (list, tuple)) and len(xy_plot_indices) >= 2:
            self.x_plot_index = xy_plot_indices[0]
            self.y_plot_index = xy_plot_indices[1]
        else:
            self.x_plot_index = 0
            self.y_plot_index = 1

        # sets excel_formats attribute
        self._create_excel_writer_formats(excel_writer_formats)


    def __str__(self):
        return f'{self.__module__}.{self.__class__.__name__} {self.name}'


    def __repr__(self):
        return f'<{str(self)}>'


    def _create_excel_writer_formats(self, format_kwargs):
        """
        Sets the excel_formats attribute for the DataSource.

        Parameters
        ----------
        format_kwargs : dict
            The input dictionary to override the default formats.

        """

        self.excel_formats = {
            'header_even': {
                'font': Font(size=12, bold=True),
                'fill': PatternFill(
                    fill_type='solid', start_color='F9B381', end_color='F9B381'
                ),
                'border': Border(bottom=Side(style='thin')),
                'alignment': Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )
            },
            'header_odd': {
                'font': Font(size=12, bold=True),
                'fill': PatternFill(
                    fill_type='solid', start_color='73A2DB', end_color='73A2DB'
                ),
                'border': Border(bottom=Side(style='thin')),
                'alignment': Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )
            },
            'subheader_even': {
                'font': Font(bold=True),
                'fill': PatternFill(
                    fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'
                ),
                'border': Border(bottom=Side(style='thin')),
                'alignment': Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )
            },
            'subheader_odd': {
                'font': Font(bold=True),
                'fill': PatternFill(
                    fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'
                ),
                'border': Border(bottom=Side(style='thin')),
                'alignment': Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )
            },
            'columns_even': {
                'fill': PatternFill(
                    fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'
                ),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'number_format': '0.00',
            },
            'columns_odd': {
                'fill': PatternFill(
                    fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'
                ),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'number_format': '0.00',
            },
            **utils.DEFAULT_FITTING_FORMATS
        }

        if format_kwargs is not None:
            for key in self.excel_formats:
                if key in format_kwargs:
                    self.excel_formats[key].update(format_kwargs[key])


    def _validate_target_columns(self):
        """
        Ensures that the target columns and function names for each Function are valid.

        Raises
        ------
        ValueError
            Raises ValueError if the target column for a Function does not exist, or if
            two Function objects have the same name. If the target column does not
            exist, it can either be due to that target not existing, or that the
            functions are calculated in the wrong order.

        """

        unique_keys = [*self.unique_variables]

        for function in (self.separation_functions + self.calculation_functions
                         + self.sample_summary_functions
                         + self.dataset_summary_functions):
            # ensure function names are unique
            if function.name in unique_keys:
                raise ValueError((
                    f'The name "{function.name}" is associated with two different '
                    f'Function objects in the DataSource "{self.name}", which is not allowed.'
                ))
            # ensure targets exist
            for target in function.target_columns:
                if target not in unique_keys:
                    raise ValueError((
                        f'"{target}" is not an available target for Function '
                        f'"{function.name}". Check that the function order is correct.'
                    ))
            # ensure columns exist if function modifies columns
            if (not isinstance(function, SeparationFunction)
                    and not isinstance(function.added_columns, int)):

                for target in function.added_columns:
                    if target not in unique_keys:
                        raise ValueError((
                            f'"{target}" is not an available column for Function '\
                            f'"{function.name}" to modify. Check that the function order is correct.'
                        ))
            # ensure summary functions either add columns or modify other summary columns
            if (isinstance(function, SummaryFunction)
                    and not isinstance(function.added_columns, int)):

                if function.sample_summary:
                    sum_funcs = [function.name for function in self.sample_summary_functions]
                else:
                    sum_funcs = [function.name for function in self.dataset_summary_functions]

                if any(column not in sum_funcs for column in function.added_columns):
                    raise ValueError((
                        f'Error with "{function.name}". SummaryFunctions can only modify '
                        'other SummaryFunction columns.'
                    ))

            unique_keys.append(function.name)


    def _create_references(self, dataset, import_values):
        """Creates a dictionary to reference the column indices."""

        references = [[] for sample in dataset]
        for i, sample in enumerate(dataset):
            for j, dataframe in enumerate(sample):
                reference = {
                    variable: [import_values[i][j][f'index_{variable}']] for variable in self.unique_variables
                }
                start_index = len(dataframe.columns)

                for function in self.calculation_functions:
                    if isinstance(function.added_columns, int):
                        end_index = start_index + function.added_columns
                        reference[function.name] = list(range(start_index, end_index))

                        for num in range(start_index, end_index):
                            dataframe[num] = pd.Series(np.nan, dtype=np.float32)

                        start_index = end_index
                    else:
                        reference[function.name] = []
                        for target in function.added_columns:
                            reference[function.name].extend(reference[target])

                references[i].append(reference)

        return references


    def _add_summary_dataframes(self, dataset, references):
        """Adds the dataframes for summary functions and creates references for them."""

        if self.sample_summary_functions:
            for reference in references:
                reference.append({})

            for i, sample in enumerate(dataset):
                start_index = 0
                data = {}
                for function in self.sample_summary_functions:
                    if isinstance(function.added_columns, int):
                        end_index = start_index + function.added_columns
                        references[i][-1][function.name] = list(
                            range(start_index, end_index)
                        )
                        for num in range(start_index, end_index):
                            data[num] = np.nan
                        start_index = end_index

                    else:
                        references[i][-1][function.name] = []
                        for target in function.added_columns:
                            references[i][-1][function.name].extend(
                                references[i][-1][target]
                            )

                sample.append(pd.DataFrame(data, index=[0], dtype=np.float32))

        if self.dataset_summary_functions:
            references.append([{}])
            start_index = 0
            data = {}
            for function in self.dataset_summary_functions:
                if isinstance(function.added_columns, int):
                    end_index = start_index + function.added_columns
                    references[-1][-1][function.name] = list(
                        range(start_index, end_index)
                    )
                    for num in range(start_index, end_index):
                        data[num] = np.nan
                    start_index = end_index

                else:
                    references[-1][-1][function.name] = []
                    for target in function.added_columns:
                        references[-1][-1][function.name].extend(
                            references[-1][-1][target]
                        )

            dataset.append([pd.DataFrame(data, index=[0], dtype=np.float32)])


    def _merge_references(self, dataframes, references):
        """
        Merges all the references for the merged dataframe.

        """

        functions = (self.calculation_functions
                     + self.sample_summary_functions
                     + self.dataset_summary_functions)

        merged_references = []
        for i, dataset in enumerate(dataframes):
            start_index = 0
            merged_reference = {
                key: [] for key in (self.unique_variables
                                    + [func.name for func in functions])
            }
            for j, sample in enumerate(dataset):
                for key in merged_reference:
                    merged_reference[key].append([])
                for k, entry in enumerate(sample):
                    for key in references[i][j][k]:
                        merged_reference[key][j].extend([
                            index + start_index for index in references[i][j][k][key]
                        ])

                    start_index += len(entry.columns)

            merged_references.append(merged_reference)

        return merged_references


    def set_references(self, dataframes, import_values):
        """
        Creates a dictionary to reference the column indices for calculations.

        Also adds the necessary columns to the input dataframes for all calculations,
        creates dataframes for the SummaryCalculations, and adds spacing between
        samples and entries.

        Assigns the merged references to the references attribute.

        Parameters
        ----------
        dataframes : list
            A list of lists of lists of dataframes.
        import_values : list
            A list of lists of dictionaries containing the values used to import the data
            from files. The relevant keys are the DataSource's unique variables

        """

        # create references, add summary dataframes, and add in empty spacing columns
        references = []
        for i, dataset in enumerate(dataframes):
            reference = self._create_references(dataset, import_values[i])
            self._add_summary_dataframes(dataset, reference)
            references.append(reference)

            # add entry spacings
            for sample in dataset:
                for k, entry in enumerate(sample):
                    if k < len(sample) - 1:
                        start_index = len(entry.columns)
                        for num in range(start_index, start_index + self.entry_separation):
                            entry[num] = pd.Series(np.nan, dtype=np.float16)

                # add sample spacings
                start_index = len(sample[-1].columns)
                for num in range(start_index, start_index + self.sample_separation):
                    sample[-1][num] = pd.Series(np.nan, dtype=np.float16)

        # merges the references into one for each dataset
        self.references = self._merge_references(dataframes, references)


    def do_separation_functions(self, dataframes, import_values):
        """
        Performs the function for all SeparationFunctions.

        Parameters
        ----------
        dataframes : list
            A list of lists of lists of dataframes.
        import_values : list
            A list of lists of dictionaries containing the values used to import the data
            from files. The relevant keys are the DataSource's unique variables

        Returns
        -------
        new_dataframes : list
            The list of lists of lists of dataframes, after performing the
            separation calculations.
        new_import_values : list
            The list of lists of dictionaries containing the values used to
            import the data from files, after performing the
            separation calculations.

        """

        new_dataframes = []
        new_import_values = []
        for i, dataset in enumerate(dataframes):
            for function in self.separation_functions:
                dataset, import_values[i] = function.separate_dataframes(
                    dataset, import_values[i]
                )
            new_dataframes.append(dataset)
            new_import_values.append(import_values[i])

        return new_dataframes, new_import_values


    def merge_datasets(self, dataframes):
        """
        Merges all entries and samples into one dataframe for each dataset.

        Also sets the length attribute, which will later be used to separate each
        dataframes back into individual dataframes for each entry.

        Parameters
        ----------
        dataframes : list
            A nested list of list of lists of dataframes.

        Returns
        -------
        merged_dataframes : list
            A list of dataframes.

        """

        merged_dataframes = []
        # length of each individual entry for splitting later
        lengths = [[[] for sample in dataset] for dataset in dataframes]
        for i, dataset in enumerate(dataframes):
            for j, sample in enumerate(dataset):
                lengths[i][j] = [len(entry.columns) for entry in sample]
            # merges all dataframes in the dataset using generators
            dataset_dataframe = pd.concat(
                (pd.concat(
                    (entry for entry in sample), axis=1) for sample in dataset),
                axis=1
            )
            dataset_dataframe.columns = list(range(len(dataset_dataframe.columns)))
            merged_dataframes.append(dataset_dataframe)

        self.lengths = lengths

        return merged_dataframes


    def _do_functions(self, dataframes, index):
        """
        Performs the function for all CalculationFunctions and SummaryFunctions.

        Parameters
        ----------
        dataframes : list
            A list of dataframes, one per dataset.
        index : int
            If index is 0, will perform the Excel functions; if index is 1, will
            perform the python functions.

        Returns
        -------
        processed_dataframes : list
            The list of dataframes after processing.

        Notes
        -----
        The start row is set to self.excel_row_offset + 3 since openpyxl is 1-based
        and there are two header rows. The start column is set to
        self.excel_column_offset + 1 since openpyxl is 1-based.

        """

        functions = (self.calculation_functions + self.sample_summary_functions
                     + self.dataset_summary_functions)

        processed_dataframes = []
        for i, dataset in enumerate(dataframes):
            for function in functions:
                dataset = function.do_function(
                    dataset, self.references[i], index,
                    self.excel_column_offset + 1, self.excel_row_offset + 3
                )

            # Optimizes memory usage after calculations
            processed_dataframes.append(utils.optimize_memory(dataset, bool(index)))

        return processed_dataframes


    def do_excel_functions(self, dataframes):
        """
        Will perform the Excel function for each CalculationFunctions and SummaryFunctions.

        Convenience wrapper for the internal method _do_functions.

        Parameters
        ----------
        dataframes : list
            A list of dataframes, one for each dataset.

        Returns
        -------
        list
            The list of dataframes after processing.

        """

        return self._do_functions(dataframes, 0)


    def do_python_functions(self, dataframes):
        """
        Will perform the python function for each CalculationFunctions and SummaryFunctions.

        Convenience wrapper for the internal method _do_functions.

        Parameters
        ----------
        dataframes : list
            A list of dataframes, one for each dataset.

        Returns
        -------
        list
            The list of dataframes after processing.

        """

        return self._do_functions(dataframes, 1)


    def split_into_entries(self, merged_dataframes):
        """
        Splits the merged dataset dataframes back into dataframes for each entry.

        Parameters
        ----------
        merged_dataframes : list
            A list of dataframes. Each dataframe will be split into lists of lists
            of dataframes.

        Returns
        -------
        split_dataframes : list
            A list of lists of lists of dataframes, corresponding to entries
            and samples within each dataset.

        """

        sample_lengths = [
            np.cumsum(
                [np.sum(sample) for sample in dataset]
            ) for dataset in self.lengths
        ]

        split_dataframes = [[[] for sample in dataset] for dataset in self.lengths]
        dataset_dtypes = []
        for i, dataset in enumerate(merged_dataframes):
            dataset_dtypes.append(iter(dataset.dtypes.values))
            split_samples = np.array_split(
                dataset, sample_lengths[i], axis=1
            )[:-1]

            for j, sample in enumerate(split_samples):
                split_dataframes[i][j].extend(
                    np.array_split(
                        sample, np.cumsum(self.lengths[i][j]), axis=1)[:-1]
                )

        # renames columns back to individual indices and reassigns dtypes
        for i, dataset in enumerate(split_dataframes):
            for sample in dataset:
                for j, entry in enumerate(sample):
                    entry.columns = [*range(len(entry.columns))] #TODO later assign column names to the dataframes here
                    dtypes = {
                        col: next(dataset_dtypes[i]) for col in entry.columns
                    }
                    sample[j] = entry.astype(dtypes)

        return split_dataframes


    def create_needed_labels(self, max_df_length=None):
        """
        Calculates the necessary column labels for imported data and Functions.

        Also fills in as many labels as possible using self.column_labels.

        Parameters
        ----------
        max_df_length : int, optional
            The highest number of columns in the imported dataframes.

        Returns
        -------
        total_labels : list
            A list with four lists, containing all the needed column
            labels: index 0 is for imported data labels,
            index 1 is for CalculationFunctions labels, index 2 is for
            sample SummaryFunctions labels, and index 3 is for dataset
            SummaryFunctions labels.

        """

        total_labels = [[], [], [], []]
        total_labels[0].extend('' for _ in range(len(self.column_numbers)))

        for function in self.calculation_functions:
            if isinstance(function.added_columns, int):
                total_labels[1].extend('' for _ in range(function.added_columns))

        for function in self.sample_summary_functions:
            if isinstance(function.added_columns, int):
                total_labels[2].extend('' for _ in range(function.added_columns))

        for function in self.dataset_summary_functions:
            if isinstance(function.added_columns, int):
                total_labels[3].extend('' for _ in range(function.added_columns))

        specified_labels = iter(self.column_labels)
        fill_labels = True
        for entry in total_labels:
            if not fill_labels:
                break
            else:
                for i in range(len(entry)):
                    try:
                        entry[i] = next(specified_labels)
                    except StopIteration:
                        fill_labels = False
                        break

        # fills labels for the imported data last in case the number
        # of imported columns is different than self.column_numbers
        if max_df_length is not None:
            temp_names = total_labels[0].copy()
            total_labels[0] = ['' for _ in range(max_df_length)]
            # reassigns the column names
            for i in range(min(len(temp_names), max_df_length)):
                total_labels[0][i] = temp_names[i]

        return total_labels


    def print_column_labels_template(self):
        """
        Convenience function that will print a template for all the column headers.

        Column headers account for all of the columns imported from raw data, the
        columns added by CalculationFunctions, and the columns added by
        SummaryFunctions.

        """

        labels = self.create_needed_labels()
        label_template = list(itertools.chain.from_iterable(labels))

        print((
            f'\nImported data labels: {len(labels[0])}\n'
            f'Calculation labels: {len(labels[1])}\n'
            f'Sample summary labels: {len(labels[2])}\n'
            f'Dataset summary labels: {len(labels[3])}\n\n'
            f'label template = {label_template}'
        ))
