# -*- coding: utf-8 -*-
"""Provides a GUI to ease the use of the peak_fitting program and save the results to Excel

@author: Donald Erb
Created on May 24, 2020

"""


from collections import defaultdict
import itertools
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl.chart import Reference, Series, ScatterChart
from openpyxl.styles import NamedStyle
from openpyxl.utils.cell import get_column_letter as _get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows as _dataframe_to_rows
import pandas as pd
import PySimpleGUI as sg

from . import peak_fitting
from . import utils


def _find_peaks(dataframe, gui_values):
    """
    Finds peaks in the data according to the gui_values.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe that contains the x and y data.
    gui_values : dict
        A dictionary of values needed for finding the peaks.

    Returns
    -------
    found_peaks : list
        The list of peaks found in the data according to the
        peak finding parameters in gui_values.

    """

    headers = dataframe.columns
    x_data = dataframe[headers[gui_values['x_fit_index']]].astype(float) #TODO should change this to .loc since could be duplicate column names
    y_data = dataframe[headers[gui_values['y_fit_index']]].astype(float)
    nan_mask = (~np.isnan(x_data)) & (~np.isnan(y_data))
    x_min = max(gui_values['x_min'], min(x_data))
    x_max = min(gui_values['x_max'], max(x_data))

    additional_peaks = np.array(gui_values['peak_list'])
    additional_peaks = additional_peaks[(additional_peaks > x_min)
                                        & (additional_peaks < x_max)]

    found_peaks = peak_fitting.find_peak_centers(
        x_data[nan_mask], y_data[nan_mask], additional_peaks,
        gui_values['height'], gui_values['prominence'], x_min, x_max
    )[-1]

    return found_peaks


def _show_fit_plot(dataframe, gui_values):
    """
    Shows a plot of data with the fit range, background range, and possible peaks marked.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The dataframe that contains the x and y data.
    gui_values : dict
        A dictionary of values needed for plotting.

    """

    headers = dataframe.columns
    x_data = dataframe[headers[gui_values['x_fit_index']]].astype(float) #TODO should change this to .loc since could be duplicate column names
    y_data = dataframe[headers[gui_values['y_fit_index']]].astype(float)

    x_min = max(gui_values['x_min'], min(x_data))
    x_max = min(gui_values['x_max'], max(x_data))
    bkg_min = max(gui_values['bkg_x_min'], x_min)
    bkg_max = min(gui_values['bkg_x_max'], x_max)

    x_mid = (x_max + x_min) / 2
    bkg_mid = (bkg_max + bkg_min) / 2
    additional_peaks = np.array(gui_values['peak_list'])
    additional_peaks = additional_peaks[(additional_peaks > x_min)
                                        & (additional_peaks < x_max)]

    peaks = _find_peaks(dataframe, gui_values)

    plt.close('Fitting')
    fig, ax = plt.subplots(num='Fitting')
    ax.plot(x_data, y_data)
    ax_y = ax.get_ylim()
    y_diff = ax_y[1] - ax_y[0]

    other_peaks = False
    for peak in peaks:
        if peak not in additional_peaks:
            other_peaks = True
            found_peaks = ax.vlines(
                peak, ax_y[0] - (0.01 * y_diff), ax_y[1] + (0.03 * y_diff),
                color='green', linestyle='-.', lw=2
            )
    for peak in additional_peaks:
        user_peaks = ax.vlines(
            peak, ax_y[0] - (0.01 * y_diff), ax_y[1] + (0.03 * y_diff),
            color='blue', linestyle=':', lw=2
            )
    ax.annotate(
        '', (x_max, ax_y[1] + (0.03 * y_diff)), (x_mid, ax_y[1] + (0.03 * y_diff)),
        arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='black'),
        annotation_clip=False,
    )
    ax.annotate(
        '', (x_min, ax_y[1] + (0.03 * y_diff)), (x_mid, ax_y[1] + (0.03 * y_diff)),
        arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='black'),
        annotation_clip=False,
    )
    ax.annotate(
        'Fitting range', (x_mid, ax_y[1] + (0.063 * y_diff)), ha='center'
    )
    ax.vlines(
        x_min, ax_y[0] - (0.01 * y_diff), ax_y[1] + (0.03 * y_diff),
        color='black', linestyle='-', lw=2
    )
    ax.vlines(
        x_max, ax_y[0] - (0.01 * y_diff), ax_y[1] + (0.03 * y_diff),
        color='black', linestyle='-', lw=2
    )

    if gui_values['subtract_bkg']:
        ax.annotate(
            '', (bkg_max, ax_y[0] - (0.01 * y_diff)),
            (bkg_mid, ax_y[0] - (0.01 * y_diff)), annotation_clip=False,
            arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='red')
        )
        ax.annotate(
            '', (bkg_min, ax_y[0] - (0.01 * y_diff)),
            (bkg_mid, ax_y[0] - (0.01 * y_diff)),
            arrowprops=dict(width=1.2, headwidth=5, headlength=5, color='red'),
            annotation_clip=False
        )
        ax.annotate(
            'Background range', (bkg_mid, ax_y[0] - (0.085 * y_diff)),
            color='red', ha='center'
        )
        ax.vlines(
            bkg_min, ax_y[0] - (0.01 * y_diff), ax_y[1] + (0.03 * y_diff),
            color='red',linestyle='--', lw=2
        )
        ax.vlines(
            bkg_max, ax_y[0] - (0.01 * y_diff), ax_y[1] + (0.03 * y_diff),
            color='red', linestyle='--', lw=2
        )

    ax.set_ylim(ax_y[0] - (0.15 * y_diff), ax_y[1] + (0.15 * y_diff))

    peak_list = []
    if additional_peaks.size > 0 and other_peaks:
        peak_list = [found_peaks, user_peaks]
        label_list = ['Found peaks', 'User input peaks']
    elif additional_peaks.size > 0:
        peak_list = [user_peaks]
        label_list = ['User input peaks']
    elif peaks:
        peak_list = [found_peaks]
        label_list = ['Found peaks']

    if peak_list:
        ax.legend(
            peak_list, label_list, frameon=False, ncol=2,
            bbox_to_anchor=(0.0, 1.01, 1, 1.01), loc='lower left',
            borderaxespad=0, mode='expand'
        )

    fig.show()


#TODO split this up into several functions?
def fit_dataframe(dataframe, user_inputs=None):
    """
    Creates a GUI to select data from a dataframe for peak fitting.

    Parameters
    ----------
    dataframe : pd.DataFrame
        A pandas dataframe.
    user_inputs : dict
        Values to use as the default inputs in the GUI.

    Returns
    -------
    tuple or bool
        If peak fitting was skipped for an entry, then False is returned. Otherwise,
        a tuple is returned with the following entries:
            fit_result : list
                A list of lmfit.ModelResult objects, which give information for each
                of the fits done on the dataset.
            peaks_df : pd.DataFrame
                The dataframe containing the x and y data, the y data
                for every individual peak, the summed y data of all peaks,
                and the background, if present.
            params_df : pd.DataFrame
                The dataframe containing the value and standard error
                associated with all of the parameters in the fitting
                (eg. coefficients for the baseline, areas and sigmas for each peak).
            descriptors_df : pd.DataFrame
                The dataframe which contains some additional information about the fitting.
                Currently has the adjusted r squared, reduced chi squared, the Akaike
                information criteria, the Bayesian information criteria,
                and the minimization method used for fitting.
            gui_values : dict
                The values selected in the GUI for all of the various fields, which
                can be used to reuse the values from a past interation.

    """

    default_inputs = {
        'sample_name': 'Sample',
        'x_fit_index': '0',
        'y_fit_index': '1',
        'x_label': 'raw x data',
        'y_label': 'raw y data',
        'x_min': '-inf',
        'x_max': 'inf',
        'show_plots': True,
        'batch_fit': False,
        'peak_list': [],
        'prominence': 'inf',
        'height': '-inf',
        'model_list': [],
        'default_model': 'GaussianModel',
        'vary_Voigt': False,
        'peak_width': '',
        'center_offset': '',
        'max_sigma': 'inf',
        'min_method': 'least_squares',
        'subtract_bkg': True,
        'bkg_type': 'PolynomialModel',
        'poly_n': '0',
        'bkg_x_min': '-inf',
        'bkg_x_max': 'inf',
        'fit_residuals': False,
        'min_resid': '0.05',
        'num_resid_fits': '5',
        'automatic_peaks' : True,
        'manual_peaks': False,
        'debug': False,
        'automatic_bkg': True,
        'manual_bkg': False,
        'selected_peaks': [],
        'selected_bkg': []
    }

    if user_inputs is not None:
        default_inputs.update(user_inputs)
    # Values if using manual peak selection
    peak_list = default_inputs['selected_peaks']
    # Values if using manual background selection
    bkg_points = default_inputs['selected_bkg']
    headers = dataframe.columns

    if default_inputs['batch_fit']: #TODO need to ensure that there is data in the x and y columns, otherwise it will throw an exception
        values = default_inputs

    else:
        available_models = list(peak_fitting.peak_transformer().keys())

        if (('Voigt' in default_inputs['model_list']) or
            ('VoigtModel' == default_inputs['default_model'])) :
            disable_vary_Voigt = False
        else:
            disable_vary_Voigt = True

        automatic_layout = [
            [sg.Text('Peak x values, separated by commas (leave blank to just use peak finding):')],
            [sg.Input(', '.join(str(val) for val in default_inputs['peak_list']),
                      key='peak_list', size=(50, 1))],
            [sg.Text('Prominence:', size=(13, 1)),
             sg.Input(default_inputs['prominence'], key='prominence', size=(5, 1))],
            [sg.Text('Minimum height:', size=(13, 1)),
             sg.Input(default_inputs['height'], key='height', size=(5, 1))],
            [sg.Text('Model list, separated by commas (leave blank to just use default model):')],
            [sg.Input(', '.join(str(val) for val in default_inputs['model_list']),
                      key='model_list', size=(50, 1), enable_events=True)]
        ]
        peak_finding_layout = sg.TabGroup(
            [
                [sg.Tab('Options', automatic_layout, key='automatic_tab',
                        visible=default_inputs['automatic_peaks']),
                 sg.Tab('Options', [
                     [sg.Text('')],
                     [sg.Button('Launch Peak Selector', enable_events=True, size=(30, 5)),
                      sg.Button('Update Peak & Model Lists', enable_events=True,
                                size=(30, 5))]
                 ], key='manual_tab', visible=default_inputs['manual_peaks'])]
            ], key='tab'
        )
        auto_bkg_layout = [
            [sg.Text('Model for fitting background:'),
             sg.Combo(['PolynomialModel','ExponentialModel'], key='bkg_type',
                      readonly=False, enable_events=True,
                      default_value=default_inputs['bkg_type'])],
            [sg.Text('Polynomial order:'),
             sg.Combo(list(range(8)), key='poly_n', readonly=True,
                      default_value=default_inputs['poly_n'])],
            [sg.Text('Min and max x values to use for fitting the background:')],
            [sg.Text('    x min:', size=(8, 1)),
             sg.Input(default_inputs['bkg_x_min'], key='bkg_x_min', size=(5, 1))],
            [sg.Text('    x max:', size=(8, 1)),
             sg.Input(default_inputs['bkg_x_max'], key='bkg_x_max', size=(5, 1))],
        ]

        bkg_layout = sg.TabGroup(
            [
                [sg.Tab('Background Options', auto_bkg_layout, key='auto_bkg_tab',
                        visible=default_inputs['automatic_bkg']),
                 sg.Tab('Background Options', [
                     [sg.Text('')],
                     [sg.Button('Launch Background Selector', key='bkg_selector',
                                enable_events=True, size=(30, 5))]
                 ], key='manual_bkg_tab', visible=default_inputs['manual_bkg'])]
            ], key='bkg_tabs'
        )
        column_layout = [
            [sg.Text('Raw Data', relief='ridge', size=(60, 1),
                     justification='center')],
            [sg.Text('Sample Name:'),
             sg.Input(default_inputs['sample_name'], key='sample_name',
                      do_not_clear=True, size=(20, 1))],
            [sg.Text('Column of x data for fitting:'),
             sg.Combo(list(range(len(headers))), size=(3, 1), readonly=True,
                      key='x_fit_index', default_value=default_inputs['x_fit_index'])],
            [sg.Text('Column of y data for fitting:'),
             sg.Combo(list(range(len(headers))), size=(3, 1), readonly=True,
                      key='y_fit_index', default_value=default_inputs['y_fit_index'])],
            [sg.Text('x data label:'),
             sg.Input(default_inputs['x_label'], key='x_label',
                      do_not_clear=True, size=(20, 1))],
            [sg.Text('y data label:'),
             sg.Input(default_inputs['y_label'], key='y_label',
                      do_not_clear=True, size=(20, 1))],
            [sg.Text('Min and max values to use for fitting:')],
            [sg.Text('    x min:', size=(8, 1)),
             sg.Input(key='x_min', do_not_clear=True, size=(5, 1),
                      default_text=default_inputs['x_min'])],
            [sg.Text('    x max:', size=(8, 1)),
             sg.Input(key='x_max', do_not_clear=True, size=(5, 1),
                      default_text=default_inputs['x_max'])],
            [sg.Text('Fitting Options', relief='ridge', size=(60, 1),
                     pad=(5, (20, 10)), justification='center')],
            [sg.Text('Default peak model:'),
             sg.Combo(available_models, key='default_model', readonly=True,
                      default_value=default_inputs['default_model'], enable_events=True)],
            [sg.Checkbox('Vary gamma parameter', key='vary_Voigt',
                          disabled=disable_vary_Voigt,
                          default=default_inputs['vary_Voigt'],
                          tooltip='if True, will allow the gamma parameter in the Voigt model'\
                          ' to be varied as an additional variable')],
            [sg.Text('Peak width:', size=(11, 1)),
             sg.Input(key='peak_width', do_not_clear=True, size=(5, 1),
                      default_text=default_inputs['peak_width'])],
            [sg.Text('Center offset:', size=(11, 1)),
             sg.Input(key='center_offset', do_not_clear=True, size=(5, 1),
                      default_text=default_inputs['center_offset'])],
            [sg.Text('Maximum sigma value:'),
             sg.Input(key='max_sigma', do_not_clear=True, size=(5, 1),
                      default_text=default_inputs['max_sigma'])],
            [sg.Text('Minimization method:'),
             sg.Combo(['least_squares','leastsq'], key='min_method', readonly=False,
                      default_value=default_inputs['min_method'])],
            [sg.Checkbox('Fit residuals', key='fit_residuals', enable_events=True,
                         default=default_inputs['fit_residuals'])],
            [sg.Text('Minimum residual height:'),
             sg.Input(default_inputs['min_resid'], key='min_resid',
                      do_not_clear=True, size=(5, 1), visible=False)],
            [sg.Text('Number of residual fits:'),
             sg.Input(default_inputs['num_resid_fits'], key='num_resid_fits',
                      do_not_clear=True, size=(5, 1), visible=False)],
            [sg.Text('')],
            [sg.Checkbox('Subtract background', key='subtract_bkg',
                         enable_events=True,
                         default=default_inputs['subtract_bkg']),
             sg.Text('('),
             sg.Radio('Automatic', 'bkg_fitting',
                      key='automatic_bkg', enable_events=True,
                      default=default_inputs['automatic_bkg']),
             sg.Radio('Manual', 'bkg_fitting',
                      key='manual_bkg', enable_events=True,
                      default=default_inputs['manual_bkg']),
             sg.Text(')')],
            [bkg_layout],
            [sg.Text('Peak Finding Options', relief='ridge', size=(60, 1),
                     pad=(5, (20, 10)), justification='center')],
            [sg.Radio('Automatic Peak Finding', 'peak_finding',
                      key='automatic_peaks', enable_events=True,
                      default=default_inputs['automatic_peaks']),
             sg.Radio('Manual Peak Finding', 'peak_finding',
                      key='manual_peaks', enable_events=True,
                      default=default_inputs['manual_peaks'],
                      pad=((50, 10), 5))],
            [peak_finding_layout],
        ]

        layout = [
            [sg.Frame('', [
                [sg.Column(column_layout, scrollable=True,
                           vertical_scroll_only=True, size=(700, 500))]
            ])],
            [sg.Checkbox('Show Plots After Fitting', key='show_plots',
                         default=default_inputs['show_plots'])],
            [sg.Checkbox('Batch Fit (will not show this window again)',
                         key='batch_fit', default=default_inputs['batch_fit'])],
            [sg.Checkbox('Debug Fitting Process', key='debug',
                         default=default_inputs['debug'])],
            [sg.Text('')],
            [sg.Button('Fit', bind_return_key=True, size=(6, 1),
                       button_color=utils.PROCEED_COLOR),
             sg.Button('Test Plot'),
             sg.Button('Show Data'),
             sg.Button('Reset to Default'),
             sg.Button('Skip Fitting')]
        ]
        validations = {
            'integers': [
                ['x_fit_index', 'x column'],
                ['y_fit_index', 'y column'],
                ['poly_n', 'polynomial order'],
                ['num_resid_fits', 'number of residual fits']
            ],
            'floats': [
                ['x_min', 'x min'],
                ['x_max', 'x max'],
                ['bkg_x_min', 'background x min'],
                ['bkg_x_max', 'background x max'],
                ['peak_width', 'peak width'],
                ['height', 'minimum height'],
                ['prominence', 'prominence'],
                ['center_offset', 'center offset'],
                ['min_resid', 'minimum residual height'],
                ['max_sigma', 'maximum sigma']
            ],
            'strings': [
                ['bkg_type', 'background model'],
                ['min_method', 'minimization method'],
                ['default_model', 'default model'],
            ],
            'user_inputs': [
                ['peak_list', 'peak x values', float, True],
                ['sample_name', 'sample name', utils.string_to_unicode, False, None],
                ['x_label', 'x label', utils.string_to_unicode, False, None],
                ['y_label', 'y label', utils.string_to_unicode, False, None],
                ['model_list', 'model list', str, True],
            ]
        }

        plot_validations = {
            'integers': validations['integers'][:2],
            'floats': validations['floats'][:4] + validations['floats'][5:7],
            'user_inputs': validations['user_inputs'][:1],
        }
        peak_selector_validations = {
            'integers': validations['integers'][:3],
            'floats': validations['floats'][:5],
            'strings': validations['strings'][:1],
        }
        bkg_selector_validations = {
            'integers': validations['integers'][:2],
        }

        window = sg.Window('Peak Fitting', layout, finalize=True)
        if default_inputs['manual_peaks']:
            window['manual_tab'].select()
            window['automatic_tab'].update(visible=False)
        if default_inputs['manual_bkg']:
            window['manual_bkg_tab'].select()
            window['automatic_bkg_tab'].update(visible=False)

        while True:
            event, values = window.read()

            if event == sg.WIN_CLOSED:
                utils.safely_close_window(window)

            elif event == 'Skip Fitting':
                skip = sg.popup_yes_no(
                    'Peak fitting will be skipped for this entry.\n\nProceed?\n',
                    title='Skip Fitting'
                )
                if skip == 'Yes':
                    window.close()
                    del window
                    return False

            elif event == 'Reset to Default':
                window.fill(default_inputs)
                plt.close('Peak Selector')
                peak_list = []

            elif event == 'Test Plot':
                if utils.validate_inputs(values, **plot_validations):
                    _show_fit_plot(dataframe, values)

            elif event == 'Show Data':
                data_window = utils.show_dataframes(dataframe)
                if data_window is not None:
                    data_window.finalize().TKroot.grab_set()
                    data_window.read(close=True)
                    data_window = None

            elif event == 'Update Peak & Model Lists':
                # updates values in the window from the peak selector plot
                centers = [[peak[0][0], np.round(peak[1][0], 2)] for peak in peak_list]
                sorted_peaks = sorted(centers, key=lambda x: x[1])
                sorted_centers = ', '.join([str(center) for model, center in sorted_peaks])
                tmp_model_list = [model for model, center in sorted_peaks]
                model_list = ', '.join(tmp_model_list)
                window['peak_list'].update(value=sorted_centers)
                window['model_list'].update(value=model_list)
                if any(model in ('VoigtModel', 'SkewedVoigtModel') for model in tmp_model_list):
                    window['vary_Voigt'].update(disabled=False)
                elif values['default_model'] not in ('VoigtModel', 'SkewedVoigtModel'):
                    window['vary_Voigt'].update(disabled=True, value=False)

            elif event == 'bkg_selector':
                plt.close('Background Selector')
                if utils.validate_inputs(values, **bkg_selector_validations):
                    x_data = dataframe[headers[values['x_fit_index']]].astype(float)
                    y_data = dataframe[headers[values['y_fit_index']]].astype(float)
                    bkg_points = peak_fitting.background_selector(x_data, y_data,
                                                                  bkg_points)

            elif event == 'Launch Peak Selector':
                plt.close('Peak Selector')
                if utils.validate_inputs(values, **peak_selector_validations):
                    try:
                        headers = dataframe.columns
                        x_data = dataframe[headers[values['x_fit_index']]].astype(float)
                        y_data = dataframe[headers[values['y_fit_index']]].astype(float)
                        x_min = values['x_min']
                        x_max = values['x_max']
                        bkg_min = values['bkg_x_min']
                        bkg_max = values['bkg_x_max']
                        subtract_bkg = values['subtract_bkg']
                        background_type = values['bkg_type']
                        bkg_min = values['bkg_x_min']
                        bkg_max = values['bkg_x_max']
                        poly_n = values['poly_n']
                        peak_width = values['peak_width']
                        default_model = values['default_model']

                        if subtract_bkg and values['manual_bkg']:
                            subtract_bkg = False
                            y_subtracted = y_data.copy()
                            if len(bkg_points) > 1:
                                points = sorted(bkg_points, key=lambda x: x[0])

                                for i in range(len(points) - 1):
                                    x_points, y_points = zip(*points[i:i + 2])
                                    coeffs = np.polyfit(x_points, y_points, 1)
                                    boundary = (x_data >= x_points[0]) & (x_data <= x_points[1])
                                    x_line = x_data[boundary]
                                    y_line = y_data[boundary]
                                    y_subtracted[boundary] = y_line - np.polyval(coeffs, x_line)
                            y_data = y_subtracted

                        domain_mask = (x_data > x_min) & (x_data < x_max)
                        peak_list = peak_fitting.peak_selector(
                            x_data[domain_mask], y_data[domain_mask], peak_list,
                            peak_width, subtract_bkg, background_type, poly_n,
                            bkg_min, bkg_max, default_model
                        )
                    except Exception as e:
                        sg.popup(f'Error creating plot:\n    {repr(e)}')

            elif event in ('automatic_peaks', 'manual_peaks'):
                if event == 'automatic_peaks':
                    window['automatic_tab'].update(visible=True)
                    window['automatic_tab'].select()
                    window['manual_tab'].update(visible=False)

                else:
                    window['automatic_tab'].update(visible=False)
                    window['manual_tab'].update(visible=True)
                    window['manual_tab'].select()

            elif event in ('automatic_bkg', 'manual_bkg'):
                if event == 'automatic_bkg':
                    window['auto_bkg_tab'].update(visible=True)
                    window['auto_bkg_tab'].select()
                    window['manual_bkg_tab'].update(visible=False)
                else:
                    window['manual_bkg_tab'].update(visible=True)
                    window['manual_bkg_tab'].select()
                    window['auto_bkg_tab'].update(visible=False)

            elif event == 'subtract_bkg':
                if values['subtract_bkg']:
                    window['bkg_type'].update(visible=True)
                    window['poly_n'].update(visible=True)
                    window['bkg_x_min'].update(visible=True)
                    window['bkg_x_max'].update(visible=True)
                    window['automatic_bkg'].update(disabled=False)
                    window['manual_bkg'].update(disabled=False)
                    window['bkg_selector'].update(disabled=False)
                else:
                    window['bkg_type'].update(visible=False, value='PolynomialModel')
                    window['poly_n'].update(visible=False, value='0')
                    window['bkg_x_min'].update(visible=False, value='-inf')
                    window['bkg_x_max'].update(visible=False, value='inf')
                    window['automatic_bkg'].update(disabled=True)
                    window['manual_bkg'].update(disabled=True)
                    window['bkg_selector'].update(disabled=True)

            elif event == 'bkg_type':
                if values['bkg_type'] == 'PolynomialModel':
                    window['poly_n'].update(visible=True)
                else:
                    window['poly_n'].update(visible=False, value='0')

            elif event in ('model_list', 'default_model'):
                tmp_model_list = [
                    entry.strip() for entry in values['model_list'].split(',') if entry
                ]
                if (any(model in ('VoigtModel', 'SkewedVoigtModel') for model in tmp_model_list)
                        or values['default_model'] in ('VoigtModel', 'SkewedVoigtModel')):
                    window['vary_Voigt'].update(disabled=False)
                else:
                    window['vary_Voigt'].update(disabled=True, value=False)

            elif event == 'fit_residuals':
                if values['fit_residuals']:
                    window['min_resid'].update(visible=True)
                    window['num_resid_fits'].update(visible=True)
                else:
                    window['min_resid'].update(visible=False, value=0.05)
                    window['num_resid_fits'].update(visible=False, value=5)

            elif event =='Fit':
                if (not plt.fignum_exists('Peak Selector')
                        and not plt.fignum_exists('Background Selector')):

                    close = utils.validate_inputs(values, **validations)
                else:
                    sg.popup('Please close the Peak and/or Background Selection plots.')
                    continue

                if close:
                    for entry in values['model_list']:
                        if entry not in available_models:
                            close = False
                            sg.popup(f'Need to correct the term "{entry}" in the model list',
                                     title='Error')
                            break
                else:
                    continue

                if close:
                    if not values['manual_peaks']:
                        if not _find_peaks(dataframe, values):
                            sg.popup(
                                ('No peaks found in fitting range. Either manually enter \n'
                                 'peak positions or change peak finding options'),
                                title='Error'
                            )
                        else:
                            break
                    else:
                        if peak_list:
                            break
                        else:
                            sg.popup(
                                'Please manually select peaks or change peak finding to automatic'
                            )

        window.close()
        del window

        plt.close('Fitting')

    x_label = values['x_label']
    y_label = values['y_label']
    x_data = dataframe[headers[values['x_fit_index']]]
    y_data = dataframe[headers[values['y_fit_index']]]
    x_min = values['x_min']
    x_max = values['x_max']
    default_model = values['default_model']
    vary_Voigt = values['vary_Voigt']
    center_offset = values['center_offset']
    min_method = values['min_method']
    subtract_bkg = values['subtract_bkg']
    background_type = values['bkg_type']
    bkg_min = values['bkg_x_min']
    bkg_max = values['bkg_x_max']
    max_sigma = values['max_sigma']
    poly_n = values['poly_n']
    fit_residuals = values['fit_residuals']
    min_resid = values['min_resid']
    num_resid_fits = values['num_resid_fits']
    debug = values['debug']

    if values['manual_peaks']:
        peaks = sorted(peak_list, key=lambda x: x[1][0])
        additional_peaks = [peak[1][0] for peak in peaks]
        model_list = [peak[0][0] for peak in peaks]
        peak_width = [peak[1][2] for peak in peaks]
        peak_heights = [peak[1][1] for peak in peaks]
        values['selected_peaks'] = peak_list
        # ensures no additional peaks are found
        height = np.inf
        prominence = np.inf

    else:
        additional_peaks = values['peak_list']
        model_list = values['model_list']
        peak_width = values['peak_width']
        peak_heights = None
        values['selected_peaks'] = []
        height = values['height']
        prominence = values['prominence']

    if subtract_bkg:
        if not values['manual_bkg']:
            values['selected_bkg'] = []
        else:
            values['selected_bkg'] = bkg_points
            subtract_bkg = False
            y_subtracted = y_data.copy()
            if len(bkg_points) > 1:
                points = sorted(bkg_points, key=lambda x: x[0])
                for i in range(len(points)-1):
                    x_points, y_points = zip(*points[i:i+2])
                    coeffs = np.polyfit(x_points, y_points, 1)
                    boundary = (x_data >= x_points[0]) & (x_data <= x_points[1])
                    x_line = x_data[boundary]
                    y_line = y_data[boundary]
                    y_subtracted[boundary] = y_line - np.polyval(coeffs, x_line)

            y_data = y_subtracted

    fitting_results = peak_fitting.peak_fitting(
        x_data, y_data, height, prominence, center_offset, peak_width, default_model,
        subtract_bkg, bkg_min, bkg_max, 0, max_sigma, min_method, x_min, x_max,
        additional_peaks, model_list, background_type, poly_n, None, vary_Voigt,
        fit_residuals, num_resid_fits, min_resid, debug, peak_heights
    )

    output_list = [fitting_results[key] for key in fitting_results]
    fit_result, individual_peaks, best_values = output_list[5:]

    # Renames amplitude to area for peaks to be more clear; lmfit has amplitude==area
    # by the way the module defines the peaks
    for result in best_values:
        for param in result:
            if 'peak' in param[0] and 'amplitude' in param[0]:
                param[0] = '_'.join([*param[0].split('_')[0:2], 'area'])

    # Creation of dataframe for best values of all peak parameters
    vals = defaultdict(dict)
    std_err = defaultdict(dict)
    for term in best_values[-1]:
        if 'peak' in term[0]:
            key = f'{term[0].split("_")[0]} {int(term[0].split("_")[1]) + 1}'
            param_key = '_'.join(term[0].split('_')[2:])
        else:
            key = term[0].split('_')[0]
            param_key = term[0].split('_')[-1]
        vals[key][param_key] = term[1]
        std_err[key][param_key] = term[2]
    vals_df = pd.DataFrame(vals).transpose()
    std_err_df = pd.DataFrame(std_err).transpose()

    df_1 = pd.DataFrame()
    for name in vals_df.columns:
        df_1[f'{name}_val'] = vals_df[name]
        df_1[f'{name}_sterr'] = std_err_df[name]
    df_1 = df_1.fillna('-')

    model_names = [component._name for component in fit_result[-1].components]
    if 'pvoigt' in model_names:
        for i, name in enumerate(model_names):
            if name == 'pvoigt':
                model_names[i] = 'pseudovoigt'
    df_0 = pd.DataFrame(model_names, columns=['model'], index=vals.keys())
    params_df = pd.concat([df_0, df_1], axis=1)

    # Creation of dataframe for peak values and x and y raw data
    x = fit_result[-1].userkws['x']
    y = fit_result[-1].data
    peaks_df = pd.DataFrame()
    peaks_df[x_label] = x
    peaks_df[y_label] = y

    bkg_term = '+ background' if subtract_bkg else ''
    bkg = individual_peaks[-1]['background_'] if subtract_bkg else 0
    for term in individual_peaks[-1]:
        if 'peak' in term:
            key = f'{term.split("_")[0]} {int(term.split("_")[1]) + 1} {bkg_term}'
            peaks_df[key] = individual_peaks[-1][term] + bkg
        else:
            key = term.split('_')[0]
            peaks_df[key] = individual_peaks[-1][term]
    peaks_df['total fit'] = fit_result[-1].best_fit

    # Creation of dataframe for descriptions of the fitting
    adj_r_sq = peak_fitting.r_squared(y, fit_result[-1].best_fit,
                                      fit_result[-1].nvarys)[1]
    red_chi_sq = fit_result[-1].redchi
    bayesian_info_criteria = fit_result[-1].bic
    akaike_info_criteria = fit_result[-1].aic

    descriptors_df = pd.DataFrame(
        [adj_r_sq, red_chi_sq, akaike_info_criteria, bayesian_info_criteria,
         min_method], index=['adjusted R\u00B2', 'reduced \u03c7\u00B2',
                             'A.I.C.',  'B.I.C.', 'minimization method']
    )

    if values['show_plots']:
        peak_fitting.plot_fit_results(x, y, fit_result, True, True)
        peak_fitting.plot_individual_peaks(
            x, y, individual_peaks[-1], fit_result[-1],
            subtract_bkg, plot_w_background=True
        )
        plt.pause(0.01)

    return fit_result, peaks_df, params_df, descriptors_df, values


def fit_to_excel(peaks_dataframe, params_dataframe, descriptors_dataframe,
                 excel_writer, sheet_name=None, plot_excel=False):
    """
    Outputs the relevant data from peak fitting to an Excel file.

    Parameters
    ----------
    peaks_dataframe : pd.DataFrame
        The dataframe containing the x and y data, the y data
        for every individual peak, the summed y data of all peaks,
        and the background, if present.
    params_dataframe : pd.DataFrame
        The dataframe containing the value and standard error
        associated with all of the parameters in the fitting
        (eg. coefficients for the baseline, areas and sigmas for each peak).
    descriptors_dataframe : pd.DataFrame
        The dataframe which contains some additional information about the fitting.
        Currently has the adjusted r squared, reduced chi squared, the Akaike
        information criteria, the Bayesian information criteria, and the minimization
        method used for fitting.
    excel_writer : pd.ExcelWriter
        The pandas ExcelWriter object that contains all of the
        information about the Excel file being created.
    sheet_name: str, optional
        The Excel sheet name.
    plot_excel : bool, optional
        If True, will create a simple plot in Excel that plots the raw x and
        y data, the data for each peak, the total of all peaks, and
        the background if it is present.

    """

    # Ensures that the sheet name is unique so it does not overwrite data;
    # not needed for openpyxl, but just a precaution
    current_sheets = [sheet.title.lower() for sheet in excel_writer.book.worksheets]
    if sheet_name is not None:
        sheet_name = utils.string_to_unicode(sheet_name)
        sheet_base = sheet_name
    else:
        sheet_base = 'Sheet'
        sheet_name = 'Sheet_1'
    num = 1
    while True:
        if sheet_name.lower() not in current_sheets:
            break
        else:
            sheet_name = f'{sheet_base}_{num}'
            num += 1

    param_names = dict.fromkeys([
        '',
        *[name.replace('_sterr', '').replace('_val', '') for name in params_dataframe.columns]
    ])
    total_width = (len(peaks_dataframe.columns) + len(params_dataframe.columns)
                   + len(descriptors_dataframe.columns) + 1)

    # Easier to just write params and descriptors using pandas rather than using
    # openpyxl; will not cost significant time since there are only a few cells
    params_dataframe.to_excel(
        excel_writer, sheet_name=sheet_name, startrow=3,
        startcol=len(peaks_dataframe.columns), header=False, index=True
    )
    descriptors_dataframe.to_excel(
        excel_writer, sheet_name=sheet_name, startrow=1,
        startcol=len(peaks_dataframe.columns) + len(params_dataframe.columns) + 1,
        header=False, index=True
    )
    worksheet = excel_writer.book[sheet_name]

    # Write and format all headers
    headers = (
        {'name': 'Values', 'start': 1, 'end': len(peaks_dataframe.columns)},
        {'name': 'Peak Parameters', 'start': len(peaks_dataframe.columns) + 1,
         'end': len(peaks_dataframe.columns) + len(params_dataframe.columns) + 1},
        {'name': 'Fit Description',
         'start': len(peaks_dataframe.columns) + len(params_dataframe.columns) + 2,
         'end': total_width + 1}
    )
    suffix = itertools.cycle(['even', 'odd'])
    for header in headers:
        cell = worksheet.cell(row=1, column=header['start'], value=header['name'])
        cell.style = 'fitting_header_' + next(suffix)
        worksheet.merge_cells(
            start_row=1, start_column=header['start'], end_row=1, end_column=header['end']
        )

    # Subheaders for peaks_dataframe
    cell = worksheet.cell(row=2, column=1, value='Raw Data')
    cell.style = 'fitting_header_odd'
    cell = worksheet.cell(row=2, column=3, value='Fit Data')
    cell.style = 'fitting_header_even'
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    worksheet.merge_cells(start_row=2, start_column=3, end_row=2,
                          end_column=len(peaks_dataframe.columns))

    # Formatting for peaks_dataframe
    suffix = itertools.cycle(['even', 'odd'])
    for i, peak_name in enumerate(peaks_dataframe.columns, 1):
        cell = worksheet.cell(row=3, column=i, value=peak_name)
        cell.style = 'fitting_subheader_' + next(suffix)

    rows = _dataframe_to_rows(peaks_dataframe, index=False, header=False)
    for row_index, row in enumerate(rows, 4):
        suffix = itertools.cycle(['even', 'odd'])
        for column_index, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            cell.style = 'fitting_columns_' + next(suffix)

    # Formatting for params_dataframe
    for index, subheader in enumerate(param_names):
        style_suffix = next(suffix)

        if index < 2:
            prefix = 'fitting_columns_' if index == 0 else 'fitting_subheader_'
            column = len(peaks_dataframe.columns) + 1 + index
            cell = worksheet.cell(row=2, column=column, value=subheader)
            cell.style = prefix + style_suffix
            worksheet.merge_cells(
                start_row=2, start_column=column, end_row=3, end_column=column
            )
            prefix = 'fitting_descriptors_' if index == 0 else 'fitting_columns_'
            for row in range(len(params_dataframe)):
                cell = worksheet.cell(row=4 + row, column=column)
                cell.style = prefix + style_suffix
        else:
            column = len(peaks_dataframe.columns) + 1 + (2 * (index - 1))
            cell = worksheet.cell(row=2, column=column, value=subheader)
            cell.style = 'fitting_subheader_' + style_suffix
            worksheet.merge_cells(
                start_row=2, start_column=column, end_row=2, end_column=column + 1
            )
            cell = worksheet.cell(row=3, column=column, value='value')
            cell.style = 'fitting_subheader_' + style_suffix
            cell = worksheet.cell(row=3, column=column + 1, value='standard error')
            cell.style = 'fitting_subheader_' + style_suffix

            for row in range(len(params_dataframe)):
                cell = worksheet.cell(row=4 + row, column=column)
                cell.style = 'fitting_columns_' + style_suffix
                cell = worksheet.cell(row=4 + row, column=column + 1)
                cell.style = 'fitting_columns_' + style_suffix

    # Formatting for descriptors_dataframe
    for column in range(2):
        style = 'fitting_descriptors_' + next(suffix)
        for row in range(len(descriptors_dataframe)):
            cell = worksheet.cell(
                row=2 + row,
                column=column + len(peaks_dataframe.columns) + len(params_dataframe.columns) + 2
            )
            cell.style = style

    # Adjust column and row dimensions
    worksheet.row_dimensions[1].height = 18
    for column in range(1, len(peaks_dataframe.columns) + len(params_dataframe.columns) + 2):
        worksheet.column_dimensions[_get_column_letter(column)].width = 12.5
    for column in range(len(peaks_dataframe.columns) + len(params_dataframe.columns) + 2, total_width + 2):
        worksheet.column_dimensions[_get_column_letter(column)].width = 20

    if plot_excel:
        axis_attributes = {
            'x_axis': {
                'title': peaks_dataframe.columns[0],
                'crosses': 'min'
            },
            'y_axis': {
                'title': peaks_dataframe.columns[1],
                'crosses': 'min'
            }
        }
        chart = ScatterChart()
        for axis, attribute in axis_attributes.items():
            for axis_attribute, value in attribute.items():
                setattr(getattr(chart, axis), axis_attribute, value)

        for i, peak_name in enumerate(peaks_dataframe.columns[1:], 2):
            legend_name = ' '.join(peak_name.split(' ')[0:2]) if i !=2 else 'raw data'
            chart.append(
                Series(
                    Reference(worksheet, i, 4, i, len(peaks_dataframe) + 3),
                    xvalues=Reference(worksheet, 1, 4, 1, len(peaks_dataframe) + 3),
                    title=legend_name
                )
            )

        worksheet.add_chart(chart, 'D8')


def launch_peak_fitting_gui(dataframe=None, gui_values=None, excel_writer=None,
                            save_excel=True, plot_excel=True, mpl_changes=None,
                            save_when_done=True):
    """
    Convenience function to fit dataframe(s) and write their results to Excel.

    Parameters
    ----------
    dataframe : pd.DataFrame or list/tuple, optional
        The dataframe or list/tuple of dataframes to fit.
    gui_values : dict, optional
        A dictionary containing the default gui values to pass to fit_dataframe.
    excel_writer : pd.ExcelWriter, optional
        The Excel writer used to save the results to Excel.
    save_excel : bool, optional
        If True (default), then the fit results will be saved to an Excel file.
    plot_excel : bool, optional
        If True (default), then the fit results will be plotted in the
        Excel file (if saving).
    mpl_changes : dict, optional
        A dictionary of changes to apply to matplotlib's rcParams file.
    save_when_done : bool, optional
        If True (default), then the Excel file will be saved once all dataframes
        are fit.

    Returns
    -------
    fit_results : list
        A list of lists of lmfit.ModelResult objects, which give information
        for each of the fits done on the dataframes.
    gui_values : dict, optional
        A dictionary containing the default gui values to pass to fit_dataframe.
    proceed : bool
        True if the fitting gui was not exited from prematurely, otherwise,
        the value is False. Useful when calling this function from an
        outside function that needs to know whether to continue doing
        peak fitting.

    """
    #TODO check that mpl backend is interactive; if not, try to switch to tkAgg
    rc_params = mpl_changes.copy() if mpl_changes is not None else {}
    # Correctly scales the dpi to match the desired dpi.
    dpi = rc_params.get('figure.dpi', plt.rcParams['figure.dpi'])
    dpi_scale = utils.get_dpi_correction(dpi)
    rc_params.update({
        'interactive': False,
        'figure.constrained_layout.use': False,
        'figure.autolayout': True,
        'figure.dpi': dpi_scale * dpi
    })

    if dataframe is not None:
        if isinstance(dataframe, pd.DataFrame):
            fit_dataframes = [dataframe]
        else:
            fit_dataframes = dataframe
    else:
        fit_dataframes = utils.open_multiple_files()

    if not save_excel or excel_writer is not None:
        writer = excel_writer
    else:
        layout = [
            [sg.Text('File name for peak fitting results')],
            [sg.Input('', key='file', size=(20, 1),
                      disabled=True, text_color='black'),
            sg.FileSaveAs(file_types=(("Excel Workbook (xlsx)", "*.xlsx"),),
                          key='browse', target='file')],
            [sg.Text('')],
            [sg.Button('Skip Saving'),
             sg.Button('Submit', bind_return_key=True,
                       button_color=utils.PROCEED_COLOR),
             sg.Check('New File', key='new_file')]
        ]

        window = sg.Window('File Selection', layout)
        while True:
            event, values = window.read()
            if event == sg.WIN_CLOSED:
                utils.safely_close_window(window)
            elif event == 'Skip Saving':
                save_excel = False
                writer = None
                break
            elif event == 'Submit':
                if utils.validate_inputs(values, strings=[['file', 'Excel file']]):
                    break

        window.close()
        del window

        if save_excel:
            file_path = Path(values['file'])
            if not file_path.suffix.lower() or file_path.suffix.lower() != '.xlsx':
                values['file'] = str(Path(file_path.parent, file_path.stem + '.xlsx'))

            if not values['new_file'] and Path(values['file']).exists():
                mode = 'a'
            else:
                mode = 'w'

            writer = pd.ExcelWriter(values['file'], engine='openpyxl', mode=mode)

    # Formatting styles for the Excel workbook
    for style, kwargs in utils.DEFAULT_FITTING_FORMATS.items():
        try:
            writer.book.add_named_style(NamedStyle(style, **kwargs))
        except AttributeError: # writer is None
            break
        except ValueError: # Style already exists in the workbook
            pass

    fit_results = []
    proceed = True
    for dataframe in fit_dataframes:
        try:
            with plt.rc_context(rc_params):
                fit_output = fit_dataframe(dataframe, gui_values)

        except (utils.WindowCloseError, KeyboardInterrupt):
            proceed = False
            break

        if not fit_output: # Fitting was skipped for the data entry
            fit_results.append(None)
        else:
            fit_results.append(fit_output[0])
            peak_df = fit_output[1]
            params_df = fit_output[2]
            descriptors_df = fit_output[3]
            gui_values = fit_output[4]

            if save_excel:
                fit_to_excel(peak_df, params_df, descriptors_df,
                             writer, gui_values['sample_name'], plot_excel)

    if save_excel and save_when_done and not all(entry is None for entry in fit_results):
        utils.save_excel_file(writer)

    return fit_results, gui_values, proceed
